"""Semantic workbook, cell, formula, and pattern diff tests."""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from tabulint.diff.workbook_diff import WorkbookDiffer
from tabulint.models import CellKind, RiskLevel, RuleSpec, RuleStatus, TabulintConfig
from tabulint.parser.workbook import WorkbookParser
from tabulint.risk.scorer import RiskScorer
from tabulint.rules.engine import RuleEngine
from tests.conftest import WorkbookFactory, add_vba_project


def _before_semantic_diff(workbook: Workbook) -> None:
    data = workbook.active
    data.title = "Data"
    data["A1"] = 10
    data["A2"] = "before"
    data["A3"] = date(2025, 1, 1)
    data["A4"] = "=SUM(B1:B3)"
    data["A5"] = "=B1*2"
    data["A6"] = 7
    data["A7"] = "=B2*2"
    data["A8"] = "=B3*2"
    data["B1"] = 1
    data["B2"] = 2
    data["B3"] = 3
    workbook.create_sheet("Old Sheet")["A1"] = "remove me"
    legacy_hidden = workbook.create_sheet("Legacy Hidden")
    legacy_hidden.sheet_state = "hidden"
    legacy_hidden["A1"] = "remove me too"
    workbook.create_sheet("Audit")["A1"] = "visible"


def _after_semantic_diff(workbook: Workbook) -> None:
    data = workbook.active
    data.title = "Data"
    data["A1"] = 11
    data["A2"] = "after"
    data["A3"] = date(2026, 1, 1)
    data["A4"] = "=SUM(B1:B2)"
    data["A5"] = 42
    data["A6"] = "=B1*3"
    data["A7"] = "manual"
    data["A8"] = None
    data["B1"] = 1
    data["B2"] = 2
    data["B3"] = 3
    data["C1"] = "='[Source.xlsx]Sheet1'!A1"
    workbook.create_sheet("New Sheet")["A1"] = "new"
    hidden = workbook.create_sheet("Hidden Addition")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "review"
    audit = workbook.create_sheet("Audit")
    audit.sheet_state = "hidden"
    audit["A1"] = "visible"


def test_workbook_diff_classifies_values_formulas_and_structure(
    workbook_factory: WorkbookFactory,
) -> None:
    before_path = workbook_factory("before.xlsx", _before_semantic_diff)
    after_path = workbook_factory("after.xlsm", _after_semantic_diff)
    add_vba_project(after_path)
    parser = WorkbookParser()

    structure, cells, formulas = WorkbookDiffer().compare(
        parser.parse(before_path),
        parser.parse(after_path),
    )

    cell_types = {change.location: change.change_type for change in cells}
    assert cell_types["Data!A1"] == "numeric_value_changed"
    assert cell_types["Data!A2"] == "text_changed"
    assert cell_types["Data!A3"] == "date_changed"
    assert cell_types["Data!A4"] == "formula_changed"
    assert cell_types["Data!A5"] == "formula_overwritten"
    assert cell_types["Data!A6"] == "fixed_value_to_formula"
    assert cell_types["Data!A7"] == "formula_overwritten"
    assert cell_types["Data!A8"] == "formula_deleted"
    assert cell_types["Data!C1"] == "blank_to_formula"

    formula_by_location = {change.location: change for change in formulas}
    reduced = formula_by_location["Data!A4"]
    assert reduced.change_type == "formula_range_reduced"
    assert reduced.excluded_references == ["B3"]
    assert reduced.high_impact is True
    assert formula_by_location["Data!A5"].change_type == "formula_overwritten"
    assert formula_by_location["Data!A5"].replacement_value == 42
    assert formula_by_location["Data!A5"].replacement_kind is CellKind.NUMBER
    assert formula_by_location["Data!A5"].manual_review_recommendation
    assert formula_by_location["Data!A6"].change_type == "formula_added"
    assert formula_by_location["Data!A7"].after_formula is None
    assert formula_by_location["Data!A7"].replacement_value == "manual"
    assert formula_by_location["Data!A7"].replacement_kind is CellKind.TEXT
    assert formula_by_location["Data!A8"].change_type == "formula_deleted"
    assert formula_by_location["Data!A8"].replacement_value is None
    assert formula_by_location["Data!A8"].replacement_kind is CellKind.BLANK
    assert formula_by_location["Data!A8"].manual_review_recommendation

    structure_types = {change.change_type for change in structure}
    assert {
        "sheet_added",
        "sheet_deleted",
        "hidden_sheet_added",
        "hidden_sheet_removed",
        "sheet_visibility_changed",
        "external_link_added",
        "macro_added",
    } <= structure_types
    external = next(change for change in structure if change.change_type == "external_link_added")
    assert external.after == "Source.xlsx"
    macro = next(change for change in structure if change.change_type == "macro_added")
    assert macro.before is False
    assert macro.after is True


def _before_sheet_metadata(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Main"
    sheet.append(["Key", "Value"])
    sheet.append(["One", 1])
    sheet.append(["Two", 2])
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["C"].hidden = True
    sheet.merge_cells("D1:E1")
    sheet["D1"] = "Merged"
    sheet.freeze_panes = "B2"
    validation = DataValidation(type="list", formula1='"A,B"')
    sheet.add_data_validation(validation)
    validation.add("B2:B3")
    sheet.add_table(Table(displayName="MetadataTable", ref="A1:B3"))
    workbook.defined_names.add(DefinedName("Scope", attr_text="'Main'!$A$2"))
    workbook.create_sheet("Second")


def _after_sheet_metadata(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Main"
    sheet.append(["Key", "Value"])
    sheet.append(["One", 1])
    sheet.append(["Two", 2])
    sheet.row_dimensions[3].hidden = True
    sheet.column_dimensions["D"].hidden = True
    sheet.merge_cells("D1:F1")
    sheet["D1"] = "Merged"
    sheet.freeze_panes = "C2"
    validation = DataValidation(type="list", formula1='"A,B,C"')
    sheet.add_data_validation(validation)
    validation.add("B2:B3")
    sheet.add_table(Table(displayName="MetadataTable", ref="A1:B2"))
    workbook.defined_names.add(DefinedName("Scope", attr_text="'Main'!$A$3"))
    workbook.create_sheet("Second", 0)


def test_workbook_diff_reports_all_sheet_metadata_changes(
    workbook_factory: WorkbookFactory,
) -> None:
    parser = WorkbookParser()
    before = parser.parse(workbook_factory("metadata-before.xlsx", _before_sheet_metadata))
    after = parser.parse(workbook_factory("metadata-after.xlsx", _after_sheet_metadata))

    structure, _, _ = WorkbookDiffer().compare(before, after)

    structure_types = {change.change_type for change in structure}
    assert {
        "hidden_rows_changed",
        "hidden_columns_changed",
        "merged_cells_changed",
        "data_validations_changed",
        "freeze_panes_changed",
        "tables_changed",
        "named_range_changed",
        "sheet_order_changed",
    } <= structure_types


def test_workbook_diff_reports_external_link_and_macro_removal(
    workbook_factory: WorkbookFactory,
) -> None:
    def before_configure(workbook: Workbook) -> None:
        workbook.active["A1"] = "='[Legacy.xlsx]Data'!A1"

    before_path = workbook_factory("removal-before.xlsm", before_configure)
    add_vba_project(before_path)
    before = WorkbookParser().parse(before_path)
    after = WorkbookParser().parse(workbook_factory("removal-after.xlsx"))

    structure, _, _ = WorkbookDiffer().compare(before, after)

    structure_types = {change.change_type for change in structure}
    assert "external_link_removed" in structure_types
    assert "macro_removed" in structure_types


def test_added_and_deleted_sheet_contents_feed_diffs_and_change_rules(
    workbook_factory: WorkbookFactory,
) -> None:
    def before_configure(workbook: Workbook) -> None:
        workbook.active.title = "Main"
        workbook.active["A1"] = "stable"
        deleted = workbook.create_sheet("Deleted")
        deleted["A1"] = 10
        deleted["B1"] = "=A1*2"

    def after_configure(workbook: Workbook) -> None:
        workbook.active.title = "Main"
        workbook.active["A1"] = "stable"
        added = workbook.create_sheet("Added")
        added["A1"] = 20
        added["B1"] = "=A1*3"

    parser = WorkbookParser()
    before = parser.parse(workbook_factory("sheet-content-before.xlsx", before_configure))
    after = parser.parse(workbook_factory("sheet-content-after.xlsx", after_configure))

    structure, cells, formulas = WorkbookDiffer().compare(before, after)

    cell_types = {change.location: change.change_type for change in cells}
    assert cell_types == {
        "Added!A1": "blank_to_number",
        "Added!B1": "blank_to_formula",
        "Deleted!A1": "cell_cleared",
        "Deleted!B1": "formula_deleted",
    }
    formula_types = {change.location: change.change_type for change in formulas}
    assert formula_types == {
        "Added!B1": "formula_added",
        "Deleted!B1": "formula_deleted",
    }

    config = TabulintConfig(
        rules=[
            RuleSpec(
                name="Only Main may change",
                type="allowed_change_range",
                ranges=["Main!A1"],
            ),
            RuleSpec(name="No changed cells", type="max_changed_cells", max=0),
        ]
    )
    results = RuleEngine().evaluate(config, before, after, structure, cells)

    assert [result.status for result in results] == [RuleStatus.FAILED, RuleStatus.FAILED]
    assert results[0].evidence["outside_allowed_count"] == 4
    assert results[1].evidence["changed_cell_count"] == 4


def test_styled_blank_is_only_a_low_risk_style_change(
    workbook_factory: WorkbookFactory,
) -> None:
    def after_configure(workbook: Workbook) -> None:
        workbook.active["D4"].font = Font(bold=True)

    parser = WorkbookParser()
    before = parser.parse(workbook_factory("styled-blank-before.xlsx"))
    after = parser.parse(workbook_factory("styled-blank-after.xlsx", after_configure))

    structure, cells, formulas = WorkbookDiffer().compare(before, after)

    assert structure == []
    assert formulas == []
    assert len(cells) == 1
    change = cells[0]
    assert change.location == "Sheet!D4"
    assert change.change_type == "style_changed"
    assert change.risk_level is RiskLevel.LOW
    assert change.before is None
    assert change.after is not None
    assert change.after.value is None

    score, level, factors = RiskScorer().score(structure, cells, formulas)
    assert score == 1
    assert level is RiskLevel.LOW
    assert [(factor.risk_type, factor.points) for factor in factors] == [
        ("style_changed", 1)
    ]


def _before_patterns(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Calc"
    for row in range(2, 5):
        sheet[f"D{row}"] = row
        sheet[f"E{row}"] = row * 2
        sheet[f"F{row}"] = f"=D{row}+E{row}"
    for column in "CDE":
        sheet[f"{column}8"] = 1
        sheet[f"{column}9"] = 2
        sheet[f"{column}10"] = f"={column}8+{column}9"


def _after_patterns(workbook: Workbook) -> None:
    _before_patterns(workbook)
    sheet = workbook["Calc"]
    sheet["F3"] = 999
    sheet["D10"] = "manual"


def test_formula_overwrite_reports_column_and_row_pattern_breaks(
    workbook_factory: WorkbookFactory,
) -> None:
    parser = WorkbookParser()
    before = parser.parse(workbook_factory("pattern-before.xlsx", _before_patterns))
    after = parser.parse(workbook_factory("pattern-after.xlsx", _after_patterns))

    _, _, formula_changes = WorkbookDiffer().compare(before, after)

    changes = {change.location: change for change in formula_changes}
    column_break = changes["Calc!F3"]
    assert column_break.change_type == "formula_overwritten"
    assert column_break.high_impact is True
    assert column_break.evidence["pattern_broken"] is True
    assert column_break.evidence["pattern_axis"] == "column"
    assert {item.split(":", 1)[0] for item in column_break.neighboring_formula_pattern} == {
        "F2",
        "F4",
    }

    row_break = changes["Calc!D10"]
    assert row_break.high_impact is True
    assert row_break.evidence["pattern_axis"] == "row"
    assert {item.split(":", 1)[0] for item in row_break.neighboring_formula_pattern} == {
        "C10",
        "E10",
    }
