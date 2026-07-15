"""Workbook parser tests using dynamically generated OOXML files."""

from __future__ import annotations

import struct
from datetime import date
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from sheetproof.exceptions import WorkbookParseError
from sheetproof.models import CellKind, FormulaCalculationStatus
from sheetproof.parser.workbook import WorkbookParser
from tests.conftest import WorkbookFactory, add_vba_project

_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_MARKUP_COMPATIBILITY_NAMESPACE = (
    "http://schemas.openxmlformats.org/markup-compatibility/2006"
)
_TEST_EXTENSION_NAMESPACE = "urn:sheetproof:test-extension"
_OTHER_EXTENSION_NAMESPACE = "urn:sheetproof:other-extension"
_VBA_RELATIONSHIP_TYPE = (
    "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
)
_VBA_CONTENT_TYPE = "application/vnd.ms-office.vbaproject"


def _replace_archive_member_bytes(
    path: Path,
    member: str,
    old: bytes | tuple[bytes, ...],
    new: bytes,
) -> None:
    with ZipFile(path) as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]

    rewritten = path.with_name(f"{path.stem}-rewritten{path.suffix}")
    replaced = False
    with ZipFile(rewritten, "w") as target:
        for info, data in entries:
            if info.filename == member:
                candidates = (old,) if isinstance(old, bytes) else old
                for candidate in candidates:
                    updated = data.replace(candidate, new, 1)
                    if updated != data:
                        replaced = True
                        data = updated
                        break
            target.writestr(info, data)
    assert replaced is True
    rewritten.replace(path)


def _replace_archive_member(path: Path, member: str, replacement: bytes) -> None:
    with ZipFile(path) as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]
        archive_comment = source.comment

    rewritten = path.with_name(f"{path.stem}-member-rewritten{path.suffix}")
    replaced = False
    with ZipFile(rewritten, "w") as target:
        target.comment = archive_comment
        for info, data in entries:
            if info.filename == member:
                data = replacement
                replaced = True
            target.writestr(info, data)
    assert replaced is True
    rewritten.replace(path)


def _relocate_first_worksheet(
    path: Path,
    *,
    old: bytes | None = None,
    new: bytes | None = None,
) -> None:
    source_name = "xl/worksheets/sheet1.xml"
    target_name = "xl/custom/moved-sheet.xml"
    with ZipFile(path) as source:
        entries = [(info, source.read(info.filename)) for info in source.infolist()]

    rewritten = path.with_name(f"{path.stem}-relocated{path.suffix}")
    moved = False
    relationship_updated = False
    content_type_updated = False
    worksheet_updated = old is None
    with ZipFile(rewritten, "w") as target:
        for info, data in entries:
            if info.filename == source_name:
                if old is not None and new is not None:
                    updated = data.replace(old, new, 1)
                    worksheet_updated = updated != data
                    data = updated
                target.writestr(target_name, data)
                moved = True
                continue
            if info.filename == "xl/_rels/workbook.xml.rels":
                updated = data.replace(
                    b"worksheets/sheet1.xml",
                    b"custom/moved-sheet.xml",
                )
                relationship_updated = updated != data
                data = updated
            elif info.filename == "[Content_Types].xml":
                updated = data.replace(
                    b"worksheets/sheet1.xml",
                    b"custom/moved-sheet.xml",
                )
                content_type_updated = updated != data
                data = updated
            target.writestr(info, data)

    assert moved is True
    assert relationship_updated is True
    assert content_type_updated is True
    assert worksheet_updated is True
    rewritten.replace(path)


def _configure_feature_workbook(workbook: Workbook) -> None:
    main = workbook.active
    main.title = "Inputs"
    main["A1"] = 42
    main["A2"] = "alpha"
    main["A3"] = True
    main["A4"] = date(2026, 7, 13)
    main["B1"] = "=SUM(A1:A1)"
    main["C1"] = "='[Rates.xlsx]FX'!$A$1"
    main["A1"].font = Font(bold=True)
    main.row_dimensions[3].hidden = True
    main.column_dimensions["D"].hidden = True
    main.merge_cells("E1:F1")
    main["E1"] = "Merged"
    main.freeze_panes = "B2"

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
    )
    validation.error = "Use a value from 0 through 100"
    validation.showErrorMessage = True
    main.add_data_validation(validation)
    validation.add("A1:A4")

    main.append([])
    main.append(["Key", "Value"])
    main.append(["One", 1])
    main.append(["Two", 2])
    main.add_table(Table(displayName="InputsTable", ref="A6:B8"))

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "secret"
    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["A1"] = "metadata"
    workbook.defined_names.add(
        DefinedName("PrimaryInput", attr_text="'Inputs'!$A$1")
    )


def test_parser_captures_cells_sheets_and_structural_metadata(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("features.xlsx", _configure_feature_workbook)

    snapshot = WorkbookParser().parse(path)

    assert [sheet.name for sheet in snapshot.sheets] == ["Inputs", "Hidden", "Very Hidden"]
    assert [sheet.state for sheet in snapshot.sheets] == ["visible", "hidden", "veryHidden"]
    assert snapshot.file.name == "features.xlsx"
    assert snapshot.file.suffix == ".xlsx"
    assert snapshot.file.sha256
    assert snapshot.has_vba is False
    assert snapshot.parsed_cell_count >= 13

    inputs = snapshot.sheet_map()["Inputs"]
    assert inputs.hidden_rows == [3]
    assert inputs.hidden_columns == ["D"]
    assert inputs.merged_cells == ["E1:F1"]
    assert inputs.freeze_panes == "B2"
    assert inputs.cells["A1"].kind is CellKind.NUMBER
    assert inputs.cells["A2"].kind is CellKind.TEXT
    assert inputs.cells["A3"].kind is CellKind.BOOLEAN
    assert inputs.cells["A4"].kind is CellKind.DATE
    assert inputs.cells["A4"].value == "2026-07-13T00:00:00"
    assert inputs.cells["B1"].formula == "=SUM(A1:A1)"
    assert inputs.cells["B1"].cached_value is None
    assert inputs.cells["B1"].calculation_status is FormulaCalculationStatus.UNCALCULATED
    assert inputs.cells["A1"].calculation_status is None
    assert inputs.cells["A1"].style.style_id != 0
    assert inputs.tables[0].name == "InputsTable"
    assert inputs.tables[0].ref == "A6:B8"

    validation = inputs.data_validations[0]
    assert validation.ranges == "A1:A4"
    assert validation.validation_type == "whole"
    assert validation.operator == "between"
    assert validation.formula1 == "0"
    assert validation.formula2 == "100"
    assert validation.allow_blank is True
    assert validation.show_error_message is True
    assert snapshot.named_ranges[0].name == "PrimaryInput"
    assert snapshot.named_ranges[0].value == "'Inputs'!$A$1"
    assert snapshot.external_links == ["Rates.xlsx"]


def test_parser_marks_available_formula_cache_explicitly(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active["A1"] = "=1+1"

    path = workbook_factory("cached-formula.xlsx", configure)
    _replace_archive_member_bytes(
        path,
        "xl/worksheets/sheet1.xml",
        (b"<f>1+1</f><v></v>", b"<f>1+1</f><v/>", b"<f>1+1</f><v />"),
        b"<f>1+1</f><v>2</v>",
    )

    formula_cell = WorkbookParser().parse(path).sheets[0].cells["A1"]

    assert formula_cell.cached_value == 2
    assert formula_cell.calculation_status is FormulaCalculationStatus.CACHED


def test_parser_detects_inert_vba_package_member(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("macro-enabled.xlsm")
    add_vba_project(path)

    snapshot = WorkbookParser().parse(path)

    assert snapshot.has_vba is True
    assert snapshot.file.has_vba is True
    assert snapshot.file.suffix == ".xlsm"
    with ZipFile(path) as archive:
        assert archive.read("xl/vbaProject.bin") == b"SheetProof test VBA marker"


def test_parser_detects_relocated_vba_project_from_opc_metadata(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("relocated-macro.xlsx")
    add_vba_project(path, part_name="xl/custom/project.bin")

    snapshot = WorkbookParser().parse(path)

    assert snapshot.has_vba is True
    assert snapshot.file.has_vba is True
    assert snapshot.file.suffix == ".xlsx"


def test_parser_accepts_vba_default_content_type(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("default-content-type.xlsm")
    add_vba_project(path, content_type_as_default=True)

    assert WorkbookParser().parse(path).has_vba is True


def test_parser_matches_vba_part_names_case_insensitively(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("case-insensitive-macro.xlsm")
    add_vba_project(
        path,
        part_name="xl/Custom/Project.BIN",
        content_type_part_name="XL/custom/project.bin",
        relationship_target="Custom/PROJECT.bin",
    )

    assert WorkbookParser().parse(path).has_vba is True


@pytest.mark.parametrize(
    "part_name",
    [
        "xl/custom/VBA%20Project.bin",
        "xl/custom/VBA%23Project.bin",
        "xl/custom/VBA%3FProject.bin",
        "xl/custom/VBA%3AProject.bin",
    ],
    ids=["space", "hash", "question", "colon"],
)
def test_parser_preserves_percent_encoded_vba_part_names(
    workbook_factory: WorkbookFactory,
    part_name: str,
) -> None:
    path = workbook_factory("percent-encoded-macro.xlsm")
    add_vba_project(path, part_name=part_name)

    assert WorkbookParser().parse(path).has_vba is True


def test_parser_does_not_infer_vba_from_conventional_filename_alone(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("unreferenced-macro.xlsm")
    with ZipFile(path, "a", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/vbaProject.bin", b"inert unreferenced bytes")

    assert WorkbookParser().parse(path).has_vba is False


@pytest.mark.parametrize(
    ("options", "message"),
    [
        ({"include_relationship": False}, "exactly one workbook relationship"),
        ({"include_content_type": False}, "exactly one workbook relationship"),
        (
            {"content_type": "application/octet-stream"},
            "exactly one workbook relationship",
        ),
        ({"relationship_target": "missing.bin"}, "missing package part"),
    ],
    ids=["content-type-only", "relationship-only", "wrong-content-type", "missing-target"],
)
def test_parser_rejects_incomplete_vba_opc_metadata(
    workbook_factory: WorkbookFactory,
    monkeypatch: pytest.MonkeyPatch,
    options: dict[str, object],
    message: str,
) -> None:
    path = workbook_factory("incomplete-macro.xlsm")
    add_vba_project(path, **options)  # type: ignore[arg-type]

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)
    with pytest.raises(WorkbookParseError, match=message):
        WorkbookParser().parse(path)


@pytest.mark.parametrize(
    ("target", "target_mode", "message"),
    [
        ("../../outside.bin", None, "unsafe target"),
        (r"custom\project.bin", None, "unsafe URI"),
        ("custom%2Fproject.bin", None, "encoded path separator"),
        ("custom/%2E%2E/project.bin", None, "encoded unreserved character"),
        ("project%ZZ.bin", None, "invalid percent escape"),
        ("https://example.test/project.bin", "External", "must be internal"),
        ("_rels/workbook.xml.rels", None, "relationships part"),
    ],
    ids=[
        "root-escape",
        "backslash",
        "encoded-slash",
        "encoded-dot-segment",
        "invalid-percent",
        "external",
        "relationship-part",
    ],
)
def test_parser_rejects_unsafe_vba_relationship_targets(
    workbook_factory: WorkbookFactory,
    monkeypatch: pytest.MonkeyPatch,
    target: str,
    target_mode: str | None,
    message: str,
) -> None:
    path = workbook_factory("unsafe-macro-target.xlsm")
    add_vba_project(path, relationship_target=target, target_mode=target_mode)

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)
    with pytest.raises(WorkbookParseError, match=message):
        WorkbookParser().parse(path)


def test_parser_applies_content_type_override_before_default(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("overridden-macro-type.xlsm")
    add_vba_project(
        path,
        content_type_as_default=True,
        override_content_type="application/octet-stream",
    )

    with pytest.raises(WorkbookParseError, match="exactly one workbook relationship"):
        WorkbookParser().parse(path)


def test_parser_rejects_multiple_vba_projects(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("multiple-macro-projects.xlsm")
    add_vba_project(path)
    add_vba_project(path, part_name="xl/custom/project.bin")

    with pytest.raises(WorkbookParseError, match="exactly one workbook relationship"):
        WorkbookParser().parse(path)


def test_parser_rejects_case_only_archive_path_duplicates(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("case-duplicate-macro.xlsm")
    add_vba_project(path)
    with ZipFile(path, "a", compression=ZIP_DEFLATED) as archive:
        archive.writestr("XL/VBAPROJECT.BIN", b"duplicate inert bytes")

    with pytest.raises(WorkbookParseError, match="unsafe or duplicate archive path"):
        WorkbookParser().parse(path)


def test_parser_rejects_case_only_content_type_override_duplicates(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("duplicate-content-type.xlsm")
    add_vba_project(path)
    with ZipFile(path) as archive:
        content_types = archive.read("[Content_Types].xml")
    root = ElementTree.fromstring(content_types)
    ElementTree.SubElement(
        root,
        f"{{{_CONTENT_TYPES_NAMESPACE}}}Override",
        PartName="/XL/VBAPROJECT.BIN",
        ContentType="application/vnd.ms-office.vbaProject",
    )
    _replace_archive_member(
        path,
        "[Content_Types].xml",
        ElementTree.tostring(root, encoding="utf-8"),
    )

    with pytest.raises(WorkbookParseError, match="duplicate override part"):
        WorkbookParser().parse(path)


def test_parser_rejects_duplicate_relationship_ids(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("duplicate-relationship-id.xlsm")
    add_vba_project(path)
    member = "xl/_rels/workbook.xml.rels"
    with ZipFile(path) as archive:
        relationships = archive.read(member)
    root = ElementTree.fromstring(relationships)
    first = next(iter(root))
    ElementTree.SubElement(root, first.tag, dict(first.attrib))
    _replace_archive_member(path, member, ElementTree.tostring(root, encoding="utf-8"))

    with pytest.raises(WorkbookParseError, match="duplicate relationship ID"):
        WorkbookParser().parse(path)


def test_parser_accepts_safe_utf16_opc_metadata(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("utf16-content-types.xlsx")
    with ZipFile(path) as archive:
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
    replacement = (
        '<?xml version="1.0" encoding="utf-16"?>' + content_types
    ).encode("utf-16")
    _replace_archive_member(path, "[Content_Types].xml", replacement)

    assert WorkbookParser().parse(path).has_vba is False


@pytest.mark.parametrize(
    ("ignorable_prefix", "accepted"),
    [("ext", True), (None, False), ("other", False)],
    ids=["declared", "undeclared", "different-namespace"],
)
def test_parser_only_ignores_declared_relationship_extensions(
    workbook_factory: WorkbookFactory,
    ignorable_prefix: str | None,
    accepted: bool,
) -> None:
    path = workbook_factory("relationship-extension.xlsx")
    member = "_rels/.rels"
    with ZipFile(path) as archive:
        relationships = archive.read(member)
    root = ElementTree.fromstring(relationships)
    if ignorable_prefix is not None:
        root.set(
            f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}Ignorable",
            ignorable_prefix,
        )
    ElementTree.SubElement(
        root,
        f"{{{_TEST_EXTENSION_NAMESPACE}}}Metadata",
    )
    ElementTree.register_namespace("mc", _MARKUP_COMPATIBILITY_NAMESPACE)
    ElementTree.register_namespace("ext", _TEST_EXTENSION_NAMESPACE)
    if ignorable_prefix == "other":
        ElementTree.register_namespace("other", _OTHER_EXTENSION_NAMESPACE)
        root.set(f"{{{_OTHER_EXTENSION_NAMESPACE}}}Marker", "true")
    _replace_archive_member(path, member, ElementTree.tostring(root, encoding="utf-8"))

    if accepted:
        assert WorkbookParser().parse(path).has_vba is False
    else:
        with pytest.raises(WorkbookParseError, match="unexpected XML element"):
            WorkbookParser().parse(path)


def test_parser_fails_closed_on_mce_wrapped_vba_metadata(
    workbook_factory: WorkbookFactory,
) -> None:
    path = workbook_factory("mce-wrapped-macro.xlsm")
    add_vba_project(path)
    members = (
        (
            "[Content_Types].xml",
            "ContentType",
            _VBA_CONTENT_TYPE,
        ),
        (
            "xl/_rels/workbook.xml.rels",
            "Type",
            _VBA_RELATIONSHIP_TYPE,
        ),
    )
    ElementTree.register_namespace("mc", _MARKUP_COMPATIBILITY_NAMESPACE)
    ElementTree.register_namespace("ext", _TEST_EXTENSION_NAMESPACE)
    for member, attribute, expected in members:
        with ZipFile(path) as archive:
            payload = archive.read(member)
        root = ElementTree.fromstring(payload)
        wrapped = next(
            element
            for element in root
            if element.attrib.get(attribute, "").casefold() == expected.casefold()
        )
        root.remove(wrapped)
        root.set(f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}Ignorable", "ext")
        root.set(
            f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}ProcessContent",
            "ext:Wrapper",
        )
        wrapper = ElementTree.SubElement(
            root,
            f"{{{_TEST_EXTENSION_NAMESPACE}}}Wrapper",
        )
        wrapper.append(wrapped)
        _replace_archive_member(path, member, ElementTree.tostring(root, encoding="utf-8"))

    with pytest.raises(WorkbookParseError, match="markup compatibility processing"):
        WorkbookParser().parse(path)


@pytest.mark.parametrize("utf16", [False, True], ids=["utf-8", "utf-16"])
def test_parser_rejects_dtd_and_entity_opc_metadata(
    workbook_factory: WorkbookFactory,
    utf16: bool,
) -> None:
    path = workbook_factory("unsafe-content-types.xlsx")
    with ZipFile(path) as archive:
        content_types = archive.read("[Content_Types].xml")
    declaration = '<!DOCTYPE Types [<!ENTITY marker "unsafe">]>'
    if utf16:
        replacement = (
            '<?xml version="1.0" encoding="utf-16"?>'
            f"{declaration}{content_types.decode('utf-8')}"
        ).encode("utf-16")
    else:
        replacement = declaration.encode() + content_types
    _replace_archive_member(path, "[Content_Types].xml", replacement)

    with pytest.raises(WorkbookParseError, match="prohibited XML declarations"):
        WorkbookParser().parse(path)


@pytest.mark.parametrize("name", ["book.xls", "book.csv", "book.txt"])
def test_parser_rejects_unsupported_file_types(tmp_path: Path, name: str) -> None:
    path = tmp_path / name
    path.write_bytes(b"not an OOXML workbook")

    with pytest.raises(WorkbookParseError, match=r"Only \.xlsx and \.xlsm"):
        WorkbookParser().parse(path)


def test_parser_reports_missing_and_invalid_workbooks(tmp_path: Path) -> None:
    with pytest.raises(WorkbookParseError, match="does not exist"):
        WorkbookParser().parse(tmp_path / "missing.xlsx")

    invalid = tmp_path / "invalid.xlsx"
    invalid.write_bytes(b"not a zip archive")
    with pytest.raises(WorkbookParseError, match="not a valid OOXML archive"):
        WorkbookParser().parse(invalid)


def test_parser_enforces_materialized_cell_limit(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet["A1"] = 1
        sheet["A2"] = 2

    path = workbook_factory("two-cells.xlsx", configure)

    with pytest.raises(WorkbookParseError, match="more than the configured 1 cells"):
        WorkbookParser(max_cells=1).parse(path)


@pytest.mark.parametrize("invalid_limit", [0, -1, True, 1.5, "2"])
def test_parser_rejects_non_positive_or_non_integer_archive_entry_limits(
    invalid_limit: object,
) -> None:
    with pytest.raises(ValueError, match="limits must be positive integers"):
        WorkbookParser(max_archive_entries=invalid_limit)  # type: ignore[arg-type]

    assert WorkbookParser(max_archive_entries=1).max_archive_entries == 1


def test_archive_entry_limit_rejects_ordinary_zip_before_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "too-many-entries.xlsx"
    with ZipFile(path, "w") as archive:
        archive.writestr("one.txt", b"one")
        archive.writestr("two.txt", b"two")
        archive.writestr("three.txt", b"three")

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)

    with pytest.raises(
        WorkbookParseError,
        match="more than the configured 2 entries",
    ):
        WorkbookParser(max_archive_entries=2).parse(path)


def test_archive_entry_limit_ignores_forged_eocd_count(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "forged-entry-count.xlsx"
    with ZipFile(path, "w") as archive:
        for index in range(11):
            archive.writestr(f"entry-{index}.txt", b"x")

    payload = bytearray(path.read_bytes())
    eocd_offset = payload.rfind(b"PK\x05\x06")
    assert eocd_offset >= 0
    struct.pack_into("<H", payload, eocd_offset + 8, 1)
    struct.pack_into("<H", payload, eocd_offset + 10, 1)
    path.write_bytes(payload)

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)
    with pytest.raises(
        WorkbookParseError,
        match="more than the configured 10 entries",
    ):
        WorkbookParser(max_archive_entries=10).parse(path)


def test_parser_rejects_non_finite_numeric_cell(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active["A1"] = 1

    path = workbook_factory("non-finite.xlsx", configure)
    _replace_archive_member_bytes(
        path,
        "xl/worksheets/sheet1.xml",
        b"<v>1</v>",
        b"<v>1e309</v>",
    )

    with pytest.raises(WorkbookParseError, match="non-finite numeric cell value"):
        WorkbookParser().parse(path)


def test_large_merged_range_is_rejected_during_archive_preflight(
    workbook_factory: WorkbookFactory,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.merge_cells("A1:Z100")
        sheet["A1"] = "large merge"

    path = workbook_factory("large-merge.xlsx", configure)

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)

    with pytest.raises(WorkbookParseError, match="configured limit is 100"):
        WorkbookParser(max_cells=10_000, max_merged_cells=100).parse(path)
    with pytest.raises(
        WorkbookParseError,
        match="more than the configured 100 cells after merged-range expansion",
    ):
        WorkbookParser(max_cells=100, max_merged_cells=10_000).parse(path)


def test_invalid_hidden_column_range_is_rejected_without_expansion(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active.column_dimensions["C"].hidden = True

    path = workbook_factory("invalid-hidden-column.xlsx", configure)
    _replace_archive_member_bytes(
        path,
        "xl/worksheets/sheet1.xml",
        b'max="3"',
        b'max="999999999"',
    )

    with pytest.raises(
        WorkbookParseError,
        match="Hidden column range 3:999999999 is outside Excel limits",
    ):
        WorkbookParser().parse(path)


def test_parser_rejects_xfe_cell_and_merged_range(
    workbook_factory: WorkbookFactory,
) -> None:
    def cell_configure(workbook: Workbook) -> None:
        workbook.active["A1"] = 1

    cell_path = workbook_factory("xfe-cell.xlsx", cell_configure)
    _replace_archive_member_bytes(
        cell_path,
        "xl/worksheets/sheet1.xml",
        b'r="A1"',
        b'r="XFE1"',
    )

    with pytest.raises(WorkbookParseError, match="out-of-bounds cell: 'XFE1'"):
        WorkbookParser().parse(cell_path)

    def merge_configure(workbook: Workbook) -> None:
        workbook.active.merge_cells("A2:B2")
        workbook.active["A2"] = "merged"

    merge_path = workbook_factory("xfe-merge.xlsx", merge_configure)
    _replace_archive_member_bytes(
        merge_path,
        "xl/worksheets/sheet1.xml",
        (
            b'<mergeCell ref="A2:B2"/>',
            b'<mergeCell ref="A2:B2" />',
            b'<mergeCell ref="A2:B2"></mergeCell>',
        ),
        b'<mergeCell ref="XFE1:XFE2"/>',
    )

    with pytest.raises(
        WorkbookParseError,
        match="Merged range 'XFE1:XFE2' falls outside Excel limits",
    ):
        WorkbookParser().parse(merge_path)


def test_relocated_worksheet_xml_still_receives_cell_and_merge_preflight(
    workbook_factory: WorkbookFactory,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def cell_configure(workbook: Workbook) -> None:
        workbook.active["A1"] = 1

    cell_path = workbook_factory("custom-location-cell.xlsx", cell_configure)
    _relocate_first_worksheet(
        cell_path,
        old=b'r="A1"',
        new=b'r="XFE1"',
    )

    def merge_configure(workbook: Workbook) -> None:
        workbook.active.merge_cells("A1:Z100")
        workbook.active["A1"] = "large merge"

    merge_path = workbook_factory("custom-location-merge.xlsx", merge_configure)
    _relocate_first_worksheet(merge_path)

    def unexpected_load(*args: object, **kwargs: object) -> None:
        raise AssertionError(f"openpyxl load should not run: {args!r} {kwargs!r}")

    monkeypatch.setattr("sheetproof.parser.workbook.load_workbook", unexpected_load)

    with pytest.raises(WorkbookParseError, match="out-of-bounds cell: 'XFE1'"):
        WorkbookParser().parse(cell_path)
    with pytest.raises(WorkbookParseError, match="configured limit is 100"):
        WorkbookParser(max_merged_cells=100).parse(merge_path)
