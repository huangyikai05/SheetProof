"""Dependency graph and bounded impact-analysis tests."""

from __future__ import annotations

from openpyxl import Workbook

from tabulint.graph.dependency_graph import DependencyGraph
from tabulint.graph.impact_analysis import DependencyAnalyzer
from tabulint.parser.workbook import WorkbookParser
from tabulint.services.review_service import ReviewService
from tests.conftest import WorkbookFactory


def _dependency_workbook(workbook: Workbook) -> None:
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = 1
    inputs["A2"] = 2
    calc = workbook.create_sheet("Calc")
    calc["B1"] = "=SUM(Inputs!A1:A2)"
    calc["C1"] = "=B1*2"
    output = workbook.create_sheet("Output")
    output["A1"] = "=Calc!C1"


def test_dependency_graph_and_cross_sheet_impact_are_traceable(
    workbook_factory: WorkbookFactory,
) -> None:
    snapshot = WorkbookParser().parse(
        workbook_factory("dependencies.xlsx", _dependency_workbook)
    )

    graph = DependencyGraph.from_snapshot(snapshot)
    impact = DependencyAnalyzer().analyze(
        snapshot,
        ["Calc!B1"],
        ["Output!A1"],
    )[0]

    assert graph.node_count == 5
    assert graph.edge_count == 4
    assert graph.direct_upstream("Calc!B1") == ["Inputs!A1", "Inputs!A2"]
    assert graph.direct_downstream("Calc!B1") == ["Calc!C1"]
    assert impact.direct_upstream == ["Inputs!A1", "Inputs!A2"]
    assert impact.direct_downstream == ["Calc!C1"]
    assert impact.downstream_cell_count == 2
    assert impact.involved_sheet_count == 3
    assert impact.critical_cells_impacted == ["Output!A1"]
    assert ["Calc!B1", "Calc!C1", "Output!A1"] in impact.path_examples
    assert impact.cycle_detected is False
    assert impact.truncated is False


def test_review_keeps_before_upstreams_when_formula_is_overwritten(
    workbook_factory: WorkbookFactory,
) -> None:
    before_path = workbook_factory("dependency-before.xlsx", _dependency_workbook)

    def overwrite_formula(workbook: Workbook) -> None:
        _dependency_workbook(workbook)
        workbook["Calc"]["B1"] = 3

    after_path = workbook_factory("dependency-after.xlsx", overwrite_formula)

    result = ReviewService().review(before_path, after_path)

    impact = next(item for item in result.dependency_impacts if item.cell == "Calc!B1")
    assert impact.direct_upstream == ["Inputs!A1", "Inputs!A2"]
    assert impact.direct_downstream == ["Calc!C1"]
    assert impact.downstream_cell_count == 2
    assert impact.involved_sheet_count == 3
    assert impact.critical_cells_impacted == []
    assert ["Calc!B1", "Calc!C1", "Output!A1"] in impact.path_examples
    assert impact.evidence["direct_upstream_before"] == [
        "Inputs!A1",
        "Inputs!A2",
    ]
    assert impact.evidence["direct_upstream_after"] == []
    assert impact.evidence["direct_upstream_sources"] == {
        "Inputs!A1": ["before"],
        "Inputs!A2": ["before"],
    }


def test_dependency_analyzer_detects_cycles(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Cycle"
        sheet["A1"] = "=B1"
        sheet["B1"] = "=A1"

    snapshot = WorkbookParser().parse(workbook_factory("cycle.xlsx", configure))

    impact = DependencyAnalyzer().analyze(snapshot, ["Cycle!A1"], [])[0]

    assert impact.cycle_detected is True
    assert impact.downstream_cell_count == 1
    assert impact.truncated is False


def test_graph_and_traversal_limits_are_explicit(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        sheet = workbook.active
        sheet.title = "Chain"
        sheet["A1"] = 1
        for row in range(2, 7):
            sheet[f"A{row}"] = f"=A{row - 1}"
        sheet["B1"] = "=SUM(C1:C10)"

    snapshot = WorkbookParser().parse(workbook_factory("limits.xlsx", configure))

    graph = DependencyGraph.from_snapshot(snapshot, max_nodes=3)
    impact = DependencyAnalyzer(max_depth=2).analyze(snapshot, ["Chain!A1"], [])[0]

    assert graph.node_count == 3
    assert graph.truncated is True
    assert impact.downstream_cell_count == 2
    assert impact.traversal_depth == 2
    assert impact.truncated is True
    assert impact.evidence["traversal_depth_limit_reached"] is True


def test_sheet_names_are_case_normalized_for_references_and_critical_cells(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        inputs = workbook.active
        inputs.title = "Inputs"
        inputs["A1"] = 5
        calc = workbook.create_sheet("Calc")
        calc["B1"] = "=inputs!A1*2"

    snapshot = WorkbookParser().parse(workbook_factory("case-normalization.xlsx", configure))

    graph = DependencyGraph.from_snapshot(snapshot)
    impact = DependencyAnalyzer().analyze(
        snapshot,
        ["inputs!A1"],
        ["CALC!B1"],
    )[0]

    assert graph.direct_upstream("Calc!B1") == ["Inputs!A1"]
    assert graph.direct_downstream("Inputs!A1") == ["Calc!B1"]
    assert "inputs!A1" not in graph.nodes
    assert impact.cell == "Inputs!A1"
    assert impact.direct_downstream == ["Calc!B1"]
    assert impact.critical_cells_impacted == ["Calc!B1"]


def test_relative_path_external_reference_is_unsupported_without_fake_local_edge(
    workbook_factory: WorkbookFactory,
) -> None:
    def configure(workbook: Workbook) -> None:
        workbook.active.title = "Local"
        workbook.active["A1"] = r"='..\data\[Budget.xlsx]Sheet1'!A1"

    snapshot = WorkbookParser().parse(workbook_factory("relative-external.xlsx", configure))

    graph = DependencyGraph.from_snapshot(snapshot)
    analysis = graph.formula_analyses["Local!A1"]

    assert snapshot.external_links == ["Budget.xlsx"]
    assert analysis.supported is False
    assert analysis.references == ()
    assert any(
        reason.startswith("unsupported_reference:")
        for reason in analysis.unsupported_reasons
    )
    assert graph.unsupported_formulas == {"Local!A1"}
    assert graph.nodes == {"Local!A1"}
    assert graph.edge_count == 0
    assert graph.direct_upstream("Local!A1") == []
