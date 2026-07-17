"""Read workbook facts without executing macros, formulas, or external content."""

from __future__ import annotations

import hashlib
import json
import math
import posixpath
import re
import struct
import zipfile
from collections.abc import Iterable
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any, cast
from urllib.parse import unquote_to_bytes, urlsplit
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from tabulint.exceptions import WorkbookParseError
from tabulint.models import (
    CellKind,
    CellSnapshot,
    DataValidationInfo,
    FileInfo,
    FormulaCalculationStatus,
    NamedRangeInfo,
    ScalarValue,
    SheetSnapshot,
    StyleSummary,
    TableInfo,
    WorkbookSnapshot,
)

_EXTERNAL_BOOK_RE = re.compile(r"\[([^\]]+\.(?:xlsx|xlsm|xlsb|xls))\]", re.IGNORECASE)
_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_WORKSHEET_RELATIONSHIP = "/worksheet"
_WORKSHEET_CONTENT_TYPE = "/worksheet+xml"
_CONTENT_TYPES_PART = "[Content_Types].xml"
_ROOT_RELATIONSHIPS_PART = "_rels/.rels"
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_MARKUP_COMPATIBILITY_NAMESPACE = (
    "http://schemas.openxmlformats.org/markup-compatibility/2006"
)
_OFFICE_DOCUMENT_RELATIONSHIP_TYPES = {
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument",
    "http://purl.oclc.org/ooxml/officeDocument/relationships/officeDocument",
}
_VBA_RELATIONSHIP_TYPE = (
    "http://schemas.microsoft.com/office/2006/relationships/vbaProject"
)
_VBA_CONTENT_TYPE = "application/vnd.ms-office.vbaproject"
_MAX_OPC_METADATA_SIZE = 2 * 1024 * 1024
_EOCD_SIGNATURE = b"PK\x05\x06"
_CENTRAL_FILE_SIGNATURE = b"PK\x01\x02"
_CENTRAL_SIGNATURE_SIGNATURE = b"PK\x05\x05"
_MAX_ZIP_COMMENT = 65_535


class WorkbookParser:
    """Parse deterministic workbook metadata with explicit resource limits."""

    def __init__(
        self,
        *,
        max_file_size: int = 100 * 1024 * 1024,
        max_uncompressed_size: int = 512 * 1024 * 1024,
        max_archive_entries: int = 10_000,
        max_cells: int = 1_000_000,
        max_merged_cells: int = 100_000,
    ) -> None:
        limits = (
            max_file_size,
            max_uncompressed_size,
            max_archive_entries,
            max_cells,
            max_merged_cells,
        )
        if any(
            isinstance(value, bool) or not isinstance(value, int) or value < 1
            for value in limits
        ):
            raise ValueError("Workbook parser limits must be positive integers")
        self.max_file_size = max_file_size
        self.max_uncompressed_size = max_uncompressed_size
        self.max_archive_entries = max_archive_entries
        self.max_cells = max_cells
        self.max_merged_cells = max_merged_cells

    def parse(self, workbook_path: str | Path) -> WorkbookSnapshot:
        path = Path(workbook_path).expanduser().resolve()
        self._validate_path(path)
        has_vba = self._validate_archive(path)

        try:
            workbook = load_workbook(
                path,
                data_only=False,
                read_only=False,
                # VBA is detected during bounded OPC preflight; do not copy its payload.
                keep_vba=False,
                keep_links=True,
            )
        except Exception as exc:
            raise WorkbookParseError(f"Unable to open workbook '{path.name}': {exc}") from exc

        try:
            cached_workbook = load_workbook(
                path,
                data_only=True,
                read_only=False,
                keep_vba=False,
                keep_links=False,
            )
        except Exception as exc:
            workbook.close()
            raise WorkbookParseError(f"Unable to open workbook '{path.name}': {exc}") from exc

        try:
            sheets: list[SheetSnapshot] = []
            parsed_cells = 0
            for index, worksheet in enumerate(workbook.worksheets):
                cached_sheet = cached_workbook[worksheet.title]
                cells: dict[str, CellSnapshot] = {}
                raw_cells: Iterable[Any] = getattr(worksheet, "_cells", {}).values()
                for cell in raw_cells:
                    if isinstance(cell, MergedCell):
                        continue
                    parsed_cells += 1
                    if parsed_cells > self.max_cells:
                        raise WorkbookParseError(
                            f"Workbook contains more than the configured {self.max_cells:,} cells"
                        )
                    cached_value = None
                    if self._is_formula(cell.value, cell.data_type):
                        cached_value = self._normalise_value(cached_sheet[cell.coordinate].value)
                    cells[cell.coordinate] = self._cell_snapshot(cell, cached_value)

                sheets.append(
                    SheetSnapshot(
                        name=worksheet.title,
                        index=index,
                        state=worksheet.sheet_state,
                        cells=cells,
                        hidden_rows=self._hidden_rows(worksheet.row_dimensions.items()),
                        hidden_columns=self._hidden_columns(worksheet.column_dimensions.values()),
                        merged_cells=sorted(str(item) for item in worksheet.merged_cells.ranges),
                        data_validations=self._data_validations(worksheet),
                        freeze_panes=(
                            str(worksheet.freeze_panes)
                            if worksheet.freeze_panes is not None
                            else None
                        ),
                        tables=sorted(
                            (
                                TableInfo(
                                    name=table.name,
                                    display_name=table.displayName,
                                    ref=table.ref,
                                    totals_row_shown=table.totalsRowShown,
                                )
                                for table in worksheet.tables.values()
                            ),
                            key=lambda item: item.name,
                        ),
                    )
                )

            external_links = self._external_links(workbook, sheets)
            digest = self._sha256(path)
            return WorkbookSnapshot(
                file=FileInfo.from_path(path, sha256=digest, has_vba=has_vba),
                sheets=sheets,
                named_ranges=self._named_ranges(workbook),
                external_links=external_links,
                has_vba=has_vba,
                parsed_cell_count=parsed_cells,
            )
        finally:
            workbook.close()
            cached_workbook.close()

    def _validate_path(self, path: Path) -> None:
        if not path.exists():
            raise WorkbookParseError(f"Workbook does not exist: {path}")
        if not path.is_file():
            raise WorkbookParseError(f"Workbook path is not a file: {path}")
        if path.suffix.lower() not in _SUPPORTED_SUFFIXES:
            raise WorkbookParseError("Only .xlsx and .xlsm workbooks are supported")
        size = path.stat().st_size
        if size > self.max_file_size:
            raise WorkbookParseError(
                f"Workbook is {size:,} bytes; maximum allowed size is {self.max_file_size:,}"
            )

    def _validate_archive(self, path: Path) -> bool:
        self._preflight_central_directory(path)
        try:
            with zipfile.ZipFile(path) as archive:
                infos = archive.infolist()
                if len(infos) > self.max_archive_entries:
                    raise WorkbookParseError(
                        "Workbook archive contains more than the configured "
                        f"{self.max_archive_entries:,} entries"
                    )
                normalised_names: set[str] = set()
                for info in infos:
                    raw_name = info.filename
                    normalised_name = self._normalise_archive_name(raw_name)
                    normalised_key = normalised_name.casefold()
                    if (
                        not normalised_name
                        or raw_name.startswith(("/", "\\"))
                        or "\\" in raw_name
                        or "//" in raw_name
                        or any(part in {".", ".."} for part in raw_name.split("/"))
                        or normalised_key in normalised_names
                    ):
                        raise WorkbookParseError(
                            f"Workbook contains an unsafe or duplicate archive path: {raw_name!r}"
                        )
                    normalised_names.add(normalised_key)
                total_uncompressed = sum(info.file_size for info in infos)
                if total_uncompressed > self.max_uncompressed_size:
                    raise WorkbookParseError(
                        "Workbook archive expands beyond the configured safety limit"
                    )
                for info in infos:
                    if info.compress_size and info.file_size / info.compress_size > 2_000:
                        raise WorkbookParseError(
                            "Workbook contains a suspiciously compressed entry"
                        )
                has_vba = self._has_vba_project(archive, infos)
                self._preflight_sheet_xml(archive, infos)
                return has_vba
        except WorkbookParseError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise WorkbookParseError(f"Workbook is not a valid OOXML archive: {exc}") from exc

    def _has_vba_project(
        self,
        archive: zipfile.ZipFile,
        infos: list[zipfile.ZipInfo],
    ) -> bool:
        """Detect a coherent, inert VBA project without reading its binary payload."""

        parts = self._regular_archive_parts(infos)
        defaults, overrides = self._content_types(archive, parts)
        vba_parts = {
            part_key
            for part_key in parts
            if self._effective_content_type(part_key, defaults, overrides)
            == _VBA_CONTENT_TYPE
        }

        workbook_part = self._workbook_part(archive, parts)
        relationship_name = self._relationship_part_name(workbook_part)
        relationship_key = self._archive_name_key(relationship_name)
        relationship_info = parts.get(relationship_key)
        relationships = (
            self._relationships(archive, relationship_info, "workbook relationships")
            if relationship_info is not None
            else []
        )
        vba_relationships = [
            relationship
            for relationship in relationships
            if relationship.attrib["Type"] == _VBA_RELATIONSHIP_TYPE
        ]

        if not vba_relationships and not vba_parts:
            return False
        if len(vba_relationships) != 1 or len(vba_parts) != 1:
            raise WorkbookParseError(
                "VBA project metadata must contain exactly one workbook relationship "
                "and one VBA content-type part"
            )

        target_key = self._resolve_internal_relationship_target(
            workbook_part,
            vba_relationships[0],
            parts,
            "VBA project relationship",
        )
        if target_key != next(iter(vba_parts)):
            raise WorkbookParseError(
                "VBA project relationship and content type identify different package parts"
            )
        return True

    def _regular_archive_parts(
        self,
        infos: list[zipfile.ZipInfo],
    ) -> dict[str, zipfile.ZipInfo]:
        return {
            self._archive_name_key(info.filename): info
            for info in infos
            if not info.is_dir()
        }

    def _content_types(
        self,
        archive: zipfile.ZipFile,
        parts: dict[str, zipfile.ZipInfo],
    ) -> tuple[dict[str, str], dict[str, str]]:
        info = parts.get(self._archive_name_key(_CONTENT_TYPES_PART))
        if info is None:
            raise WorkbookParseError("Workbook package is missing [Content_Types].xml")
        root, ignorable_namespaces = self._read_opc_xml(archive, info, "content types")
        if root.tag != f"{{{_CONTENT_TYPES_NAMESPACE}}}Types":
            raise WorkbookParseError("Workbook content types use an invalid XML namespace")

        defaults: dict[str, str] = {}
        overrides: dict[str, str] = {}
        default_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Default"
        override_tag = f"{{{_CONTENT_TYPES_NAMESPACE}}}Override"
        for element in root:
            if self._is_ignorable_extension(
                element,
                _CONTENT_TYPES_NAMESPACE,
                ignorable_namespaces,
                "content types",
            ):
                continue
            content_type = element.attrib.get("ContentType", "").strip().casefold()
            if not content_type:
                raise WorkbookParseError("Workbook content types contain an empty media type")
            if element.tag == default_tag:
                extension = element.attrib.get("Extension", "").strip().casefold()
                if (
                    not extension
                    or extension.startswith(".")
                    or any(character in extension for character in "/\\%")
                    or self._contains_control_character(extension)
                ):
                    raise WorkbookParseError(
                        "Workbook content types contain an invalid default extension"
                    )
                if extension in defaults:
                    raise WorkbookParseError(
                        "Workbook content types contain a duplicate default extension"
                    )
                defaults[extension] = content_type
            elif element.tag == override_tag:
                part_key = self._content_type_part_key(
                    element.attrib.get("PartName", "")
                )
                if part_key in overrides:
                    raise WorkbookParseError(
                        "Workbook content types contain a duplicate override part"
                    )
                overrides[part_key] = content_type
            else:
                raise WorkbookParseError(
                    "Workbook content types contain an unexpected XML element"
                )

        for part_key, content_type in overrides.items():
            if content_type == _VBA_CONTENT_TYPE and part_key not in parts:
                raise WorkbookParseError(
                    "VBA content type references a missing package part"
                )
        return defaults, overrides

    @staticmethod
    def _effective_content_type(
        part_key: str,
        defaults: dict[str, str],
        overrides: dict[str, str],
    ) -> str | None:
        override = overrides.get(part_key)
        if override is not None:
            return override
        filename = posixpath.basename(part_key)
        if "." not in filename:
            return None
        return defaults.get(filename.rsplit(".", 1)[-1].casefold())

    def _workbook_part(
        self,
        archive: zipfile.ZipFile,
        parts: dict[str, zipfile.ZipInfo],
    ) -> str:
        info = parts.get(self._archive_name_key(_ROOT_RELATIONSHIPS_PART))
        if info is None:
            raise WorkbookParseError("Workbook package is missing root relationships")
        relationships = self._relationships(archive, info, "root relationships")
        workbook_relationships = [
            relationship
            for relationship in relationships
            if relationship.attrib["Type"] in _OFFICE_DOCUMENT_RELATIONSHIP_TYPES
        ]
        if len(workbook_relationships) != 1:
            raise WorkbookParseError(
                "Workbook package must contain exactly one office document relationship"
            )
        workbook_key = self._resolve_internal_relationship_target(
            "",
            workbook_relationships[0],
            parts,
            "office document relationship",
        )
        return self._normalise_archive_name(parts[workbook_key].filename)

    def _relationships(
        self,
        archive: zipfile.ZipFile,
        info: zipfile.ZipInfo,
        label: str,
    ) -> list[ElementTree.Element]:
        root, ignorable_namespaces = self._read_opc_xml(archive, info, label)
        if root.tag != f"{{{_RELATIONSHIPS_NAMESPACE}}}Relationships":
            raise WorkbookParseError(f"Workbook {label} use an invalid XML namespace")

        relationship_tag = f"{{{_RELATIONSHIPS_NAMESPACE}}}Relationship"
        relationships: list[ElementTree.Element] = []
        relationship_ids: set[str] = set()
        for element in root:
            if element.tag != relationship_tag:
                if self._is_ignorable_extension(
                    element,
                    _RELATIONSHIPS_NAMESPACE,
                    ignorable_namespaces,
                    label,
                ):
                    continue
                raise WorkbookParseError(
                    f"Workbook {label} contain an unexpected XML element"
                )
            relationship_id = element.attrib.get("Id", "")
            relationship_type = element.attrib.get("Type", "")
            target = element.attrib.get("Target", "")
            if not relationship_id or not relationship_type or not target:
                raise WorkbookParseError(
                    f"Workbook {label} contain an incomplete relationship"
                )
            if relationship_id in relationship_ids:
                raise WorkbookParseError(
                    f"Workbook {label} contain a duplicate relationship ID"
                )
            relationship_ids.add(relationship_id)
            relationships.append(element)
        return relationships

    def _read_opc_xml(
        self,
        archive: zipfile.ZipFile,
        info: zipfile.ZipInfo,
        label: str,
    ) -> tuple[ElementTree.Element, frozenset[str]]:
        if info.file_size > _MAX_OPC_METADATA_SIZE:
            raise WorkbookParseError(
                f"Workbook {label} exceed the {_MAX_OPC_METADATA_SIZE:,}-byte safety limit"
            )
        try:
            payload = archive.read(info)
        except (NotImplementedError, OSError, RuntimeError, zipfile.BadZipFile) as exc:
            raise WorkbookParseError(f"Unable to read workbook {label} safely: {exc}") from exc
        if len(payload) > _MAX_OPC_METADATA_SIZE:
            raise WorkbookParseError(
                f"Workbook {label} exceed the {_MAX_OPC_METADATA_SIZE:,}-byte safety limit"
            )
        if self._contains_prohibited_xml_declaration(payload, label):
            raise WorkbookParseError(f"Workbook {label} contain prohibited XML declarations")
        try:
            root = ElementTree.fromstring(payload)
            namespace_map = self._root_namespace_map(payload)
        except (ElementTree.ParseError, TypeError, ValueError) as exc:
            raise WorkbookParseError(f"Unable to parse workbook {label} safely: {exc}") from exc
        ignorable_namespaces = self._ignorable_namespaces(root, namespace_map, label)
        self._reject_unsupported_markup_compatibility(root, label)
        return root, ignorable_namespaces

    @staticmethod
    def _root_namespace_map(payload: bytes) -> dict[str, str]:
        namespaces: dict[str, str] = {}
        for event, item in ElementTree.iterparse(
            BytesIO(payload),
            events=("start", "start-ns"),
        ):
            if event == "start-ns":
                prefix, namespace = cast(tuple[str, str], item)
                namespaces[prefix or ""] = namespace
            else:
                break
        return namespaces

    @staticmethod
    def _ignorable_namespaces(
        root: ElementTree.Element,
        namespace_map: dict[str, str],
        label: str,
    ) -> frozenset[str]:
        attribute = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}Ignorable"
        prefixes = root.attrib.get(attribute, "").split()
        namespaces: set[str] = set()
        for prefix in prefixes:
            namespace = namespace_map.get(prefix)
            if namespace is None:
                raise WorkbookParseError(
                    f"Workbook {label} declare an undefined ignorable XML prefix"
                )
            namespaces.add(namespace)
        return frozenset(namespaces)

    @staticmethod
    def _reject_unsupported_markup_compatibility(
        root: ElementTree.Element,
        label: str,
    ) -> None:
        alternate_content = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}AlternateContent"
        process_content = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}ProcessContent"
        must_understand = f"{{{_MARKUP_COMPATIBILITY_NAMESPACE}}}MustUnderstand"
        for element in root.iter():
            if (
                element.tag == alternate_content
                or process_content in element.attrib
                or must_understand in element.attrib
            ):
                raise WorkbookParseError(
                    f"Workbook {label} require unsupported markup compatibility processing"
                )

    @classmethod
    def _is_ignorable_extension(
        cls,
        element: ElementTree.Element,
        core_namespace: str,
        ignorable_namespaces: frozenset[str],
        label: str,
    ) -> bool:
        namespace = cls._xml_namespace(element.tag)
        if namespace is None or namespace == core_namespace:
            return False
        if namespace not in ignorable_namespaces:
            return False
        for descendant in element.iter():
            descendant_namespace = cls._xml_namespace(descendant.tag)
            if (
                descendant_namespace is None
                or descendant_namespace == core_namespace
                or descendant_namespace not in ignorable_namespaces
            ):
                raise WorkbookParseError(
                    f"Workbook {label} contain unsafe nested extension content"
                )
        return True

    @staticmethod
    def _xml_namespace(tag: str) -> str | None:
        if not tag.startswith("{") or "}" not in tag:
            return None
        return tag[1:].split("}", 1)[0]

    @staticmethod
    def _contains_prohibited_xml_declaration(payload: bytes, label: str) -> bool:
        if b"\x00" not in payload:
            folded_payload = payload.lower()
            return b"<!doctype" in folded_payload or b"<!entity" in folded_payload

        encoding: str | None = None
        if payload.startswith((b"\x00\x00\xfe\xff", b"\x00\x00\x00<")):
            encoding = "utf-32-be"
        elif payload.startswith((b"\xff\xfe\x00\x00", b"<\x00\x00\x00")):
            encoding = "utf-32-le"
        elif payload.startswith((b"\xfe\xff", b"\x00<\x00?")):
            encoding = "utf-16-be"
        elif payload.startswith((b"\xff\xfe", b"<\x00?\x00")):
            encoding = "utf-16-le"
        if encoding is None:
            raise WorkbookParseError(f"Workbook {label} uses an unsafe XML encoding")
        try:
            folded_text = payload.decode(encoding, errors="strict").casefold()
        except UnicodeDecodeError as exc:
            raise WorkbookParseError(
                f"Workbook {label} uses an invalid Unicode XML encoding"
            ) from exc
        return "<!doctype" in folded_text or "<!entity" in folded_text

    def _content_type_part_key(self, part_name: str) -> str:
        validated = self._validate_package_uri(part_name, "content-type part name")
        if not validated.startswith("/") or validated.startswith("//"):
            raise WorkbookParseError(
                "Workbook content type contains a non-absolute package part name"
            )
        relative_name = validated[1:]
        if (
            not relative_name
            or "//" in relative_name
            or any(part in {"", ".", ".."} for part in relative_name.split("/"))
        ):
            raise WorkbookParseError(
                "Workbook content type contains an unsafe package part name"
            )
        part_key = self._archive_name_key(relative_name)
        if not part_key:
            raise WorkbookParseError(
                "Workbook content type contains an unsafe package part name"
            )
        return part_key

    def _resolve_internal_relationship_target(
        self,
        source_part: str,
        relationship: ElementTree.Element,
        parts: dict[str, zipfile.ZipInfo],
        label: str,
    ) -> str:
        target_mode = relationship.attrib.get("TargetMode", "")
        if target_mode and target_mode.casefold() != "internal":
            raise WorkbookParseError(f"Workbook {label} target must be internal")
        target = self._validate_package_uri(
            relationship.attrib["Target"],
            f"{label} target",
        )
        if target.startswith("/"):
            combined = target[1:]
        else:
            combined = posixpath.join(posixpath.dirname(source_part), target)
        normalised = posixpath.normpath(combined)
        if (
            not normalised
            or normalised == "."
            or normalised.startswith(("../", "/"))
            or normalised.endswith("/")
        ):
            raise WorkbookParseError(f"Workbook {label} contains an unsafe target")
        if self._is_relationship_part_name(normalised):
            raise WorkbookParseError(f"Workbook {label} cannot target a relationships part")
        target_key = self._archive_name_key(normalised)
        if target_key not in parts:
            raise WorkbookParseError(f"Workbook {label} references a missing package part")
        return target_key

    @staticmethod
    def _validate_package_uri(value: str, label: str) -> str:
        if (
            not value
            or "\\" in value
            or value.endswith("/")
            or WorkbookParser._contains_control_character(value)
        ):
            raise WorkbookParseError(f"Workbook {label} contains an unsafe URI")
        index = 0
        while index < len(value):
            if value[index] != "%":
                index += 1
                continue
            if index + 2 >= len(value) or not all(
                character in "0123456789abcdefABCDEF"
                for character in value[index + 1 : index + 3]
            ):
                raise WorkbookParseError(f"Workbook {label} contains an invalid percent escape")
            escaped_byte = int(value[index + 1 : index + 3], 16)
            if escaped_byte in {0x00, 0x2F, 0x5C}:
                raise WorkbookParseError(
                    f"Workbook {label} contains an encoded path separator or NUL"
                )
            if (
                escaped_byte in b"-._~"
                or ord("0") <= escaped_byte <= ord("9")
                or ord("A") <= escaped_byte <= ord("Z")
                or ord("a") <= escaped_byte <= ord("z")
            ):
                raise WorkbookParseError(
                    f"Workbook {label} contains an encoded unreserved character"
                )
            index += 3
        try:
            decoded = unquote_to_bytes(value).decode("utf-8", errors="strict")
        except (UnicodeDecodeError, UnicodeEncodeError) as exc:
            raise WorkbookParseError(f"Workbook {label} is not valid UTF-8") from exc
        if (
            "\\" in decoded
            or "//" in decoded
            or decoded.endswith("/")
            or WorkbookParser._contains_control_character(decoded)
        ):
            raise WorkbookParseError(f"Workbook {label} contains an unsafe URI")
        try:
            parsed = urlsplit(value)
        except ValueError as exc:
            raise WorkbookParseError(f"Workbook {label} contains an invalid URI") from exc
        if parsed.scheme or parsed.netloc or parsed.query or parsed.fragment:
            raise WorkbookParseError(f"Workbook {label} contains an external or partial URI")
        return value

    @staticmethod
    def _contains_control_character(value: str) -> bool:
        return any(ord(character) < 32 or ord(character) == 127 for character in value)

    @staticmethod
    def _relationship_part_name(source_part: str) -> str:
        directory = posixpath.dirname(source_part)
        filename = f"{posixpath.basename(source_part)}.rels"
        return posixpath.join(directory, "_rels", filename)

    @staticmethod
    def _is_relationship_part_name(part_name: str) -> bool:
        folded_name = part_name.casefold()
        return folded_name == _ROOT_RELATIONSHIPS_PART or (
            "/_rels/" in folded_name and folded_name.endswith(".rels")
        )

    def _preflight_central_directory(self, path: Path) -> None:
        """Stream central headers before ``ZipFile`` can allocate a large entry list."""

        try:
            file_size = path.stat().st_size
            tail_size = min(file_size, _MAX_ZIP_COMMENT + 22)
            with path.open("rb") as stream:
                stream.seek(file_size - tail_size)
                tail = stream.read(tail_size)
                eocd_offset = self._find_eocd(tail)
                if eocd_offset is None:
                    raise WorkbookParseError(
                        "Workbook is not a valid OOXML archive: missing ZIP end record"
                    )

                (
                    disk_number,
                    directory_disk,
                    disk_entries,
                    stated_entries,
                    directory_size,
                    _directory_offset,
                ) = struct.unpack_from("<4H2L", tail, eocd_offset + 4)
                if disk_number or directory_disk or disk_entries != stated_entries:
                    raise WorkbookParseError("Multi-disk workbook archives are not supported")
                if (
                    stated_entries == 0xFFFF
                    or directory_size == 0xFFFFFFFF
                    or _directory_offset == 0xFFFFFFFF
                ):
                    raise WorkbookParseError("ZIP64 workbook archives are not supported")

                absolute_eocd = file_size - tail_size + eocd_offset
                directory_start = absolute_eocd - directory_size
                if directory_start < 0:
                    raise WorkbookParseError("Workbook central directory has invalid bounds")
                stream.seek(directory_start)
                remaining = directory_size
                counted_entries = 0
                while remaining:
                    if remaining < 4:
                        raise WorkbookParseError("Workbook central directory is truncated")
                    signature = stream.read(4)
                    remaining -= 4
                    if signature == _CENTRAL_FILE_SIGNATURE:
                        if remaining < 42:
                            raise WorkbookParseError(
                                "Workbook central directory header is truncated"
                            )
                        fixed = stream.read(42)
                        remaining -= 42
                        name_size, extra_size, comment_size = struct.unpack_from(
                            "<3H", fixed, 24
                        )
                        variable_size = name_size + extra_size + comment_size
                        if variable_size > remaining:
                            raise WorkbookParseError(
                                "Workbook central directory entry has invalid bounds"
                            )
                        stream.seek(variable_size, 1)
                        remaining -= variable_size
                        counted_entries += 1
                        if counted_entries > self.max_archive_entries:
                            raise WorkbookParseError(
                                "Workbook archive contains more than the configured "
                                f"{self.max_archive_entries:,} entries"
                            )
                    elif signature == _CENTRAL_SIGNATURE_SIGNATURE:
                        if remaining < 2:
                            raise WorkbookParseError(
                                "Workbook central-directory signature is truncated"
                            )
                        signature_size = struct.unpack("<H", stream.read(2))[0]
                        remaining -= 2
                        if signature_size != remaining:
                            raise WorkbookParseError(
                                "Workbook central-directory signature has invalid bounds"
                            )
                        stream.seek(signature_size, 1)
                        remaining = 0
                    else:
                        raise WorkbookParseError(
                            "Workbook central directory contains an invalid record"
                        )
                if counted_entries != stated_entries:
                    raise WorkbookParseError(
                        "Workbook central-directory entry count does not match its end record"
                    )
        except WorkbookParseError:
            raise
        except (OSError, struct.error) as exc:
            raise WorkbookParseError(f"Unable to preflight workbook archive: {exc}") from exc

    @staticmethod
    def _find_eocd(tail: bytes) -> int | None:
        search_end = len(tail)
        while search_end >= 0:
            offset = tail.rfind(_EOCD_SIGNATURE, 0, search_end)
            if offset < 0:
                return None
            if len(tail) >= offset + 22:
                comment_size = struct.unpack_from("<H", tail, offset + 20)[0]
                if offset + 22 + comment_size == len(tail):
                    return offset
            search_end = offset
        return None

    def _cell_snapshot(self, cell: Any, cached_value: ScalarValue) -> CellSnapshot:
        formula, formula_attributes = self._formula_details(cell.value, cell.data_type)
        raw_value = formula if formula is not None else self._normalise_value(cell.value)
        calculation_status = None
        if formula is not None:
            calculation_status = (
                FormulaCalculationStatus.CACHED
                if cached_value is not None
                else FormulaCalculationStatus.UNCALCULATED
            )
        return CellSnapshot(
            coordinate=cell.coordinate,
            value=raw_value,
            formula=formula,
            formula_attributes=formula_attributes,
            cached_value=cached_value,
            calculation_status=calculation_status,
            kind=self._cell_kind(cell, formula),
            data_type=str(cell.data_type),
            is_date=bool(cell.is_date),
            style=self._style_summary(cell),
        )

    @staticmethod
    def _is_formula(value: object, data_type: object) -> bool:
        return data_type == "f" or (isinstance(value, str) and value.startswith("="))

    @staticmethod
    def _formula_details(
        value: object,
        data_type: object,
    ) -> tuple[str | None, dict[str, ScalarValue]]:
        if not WorkbookParser._is_formula(value, data_type):
            return None, {}
        if isinstance(value, ArrayFormula):
            return value.text or "=UNSUPPORTED_ARRAY_FORMULA()", {
                "kind": "array",
                "ref": value.ref,
            }
        if isinstance(value, DataTableFormula):
            return "=UNSUPPORTED_DATA_TABLE()", {
                "kind": "data_table",
                "ref": value.ref,
                "ca": bool(value.ca),
                "dt2D": bool(value.dt2D),
                "dtr": bool(value.dtr),
                "r1": value.r1,
                "r2": value.r2,
                "del1": bool(value.del1),
                "del2": bool(value.del2),
            }
        return str(value), {}

    @staticmethod
    def _cell_kind(cell: Any, formula: str | None) -> CellKind:
        if formula is not None:
            return CellKind.FORMULA
        if cell.value is None:
            return CellKind.BLANK
        if cell.data_type == "e":
            return CellKind.ERROR
        if bool(cell.is_date):
            return CellKind.DATE
        if isinstance(cell.value, bool):
            return CellKind.BOOLEAN
        if isinstance(cell.value, int | float):
            return CellKind.NUMBER
        if isinstance(cell.value, str):
            return CellKind.TEXT
        return CellKind.OTHER

    @staticmethod
    def _normalise_value(value: object) -> ScalarValue:
        if isinstance(value, float) and not math.isfinite(value):
            raise WorkbookParseError("Workbook contains a non-finite numeric cell value")
        if value is None or isinstance(value, str | int | float | bool):
            return value
        if isinstance(value, datetime | date | time):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _style_summary(cell: Any) -> StyleSummary:
        border_sides = {
            side: {
                "style": getattr(getattr(cell.border, side, None), "style", None),
                "color": WorkbookParser._color_summary(
                    getattr(getattr(cell.border, side, None), "color", None)
                ),
            }
            for side in ("left", "right", "top", "bottom", "diagonal")
        }
        return StyleSummary(
            style_id=int(cell.style_id),
            number_format=str(cell.number_format),
            font=json.dumps(
                {
                    "name": cell.font.name,
                    "size": cell.font.sz,
                    "bold": cell.font.b,
                    "italic": cell.font.i,
                    "underline": cell.font.u,
                    "strike": cell.font.strike,
                    "vertical_alignment": cell.font.vertAlign,
                    "color": WorkbookParser._color_summary(cell.font.color),
                },
                sort_keys=True,
                default=str,
            ),
            fill=json.dumps(
                {
                    "type": cell.fill.fill_type,
                    "foreground": WorkbookParser._color_summary(cell.fill.fgColor),
                    "background": WorkbookParser._color_summary(cell.fill.bgColor),
                },
                sort_keys=True,
                default=str,
            ),
            border=json.dumps(
                {
                    "sides": border_sides,
                    "diagonal_up": cell.border.diagonalUp,
                    "diagonal_down": cell.border.diagonalDown,
                    "outline": cell.border.outline,
                },
                sort_keys=True,
                default=str,
            ),
            alignment=json.dumps(
                {
                    "horizontal": cell.alignment.horizontal,
                    "vertical": cell.alignment.vertical,
                    "wrap": cell.alignment.wrap_text,
                    "rotation": cell.alignment.textRotation,
                    "shrink_to_fit": cell.alignment.shrinkToFit,
                    "indent": cell.alignment.indent,
                },
                sort_keys=True,
                default=str,
            ),
            protection=json.dumps(
                {"locked": cell.protection.locked, "hidden": cell.protection.hidden},
                sort_keys=True,
                default=str,
            ),
        )

    @staticmethod
    def _color_summary(color: Any) -> dict[str, ScalarValue] | None:
        if color is None:
            return None
        color_type = getattr(color, "type", None)
        value = getattr(color, "value", None)
        tint = getattr(color, "tint", 0.0)
        return {
            "type": None if color_type is None else str(color_type),
            "value": WorkbookParser._normalise_value(value),
            "tint": WorkbookParser._normalise_value(tint),
        }

    def _preflight_sheet_xml(
        self,
        archive: zipfile.ZipFile,
        infos: list[zipfile.ZipInfo],
    ) -> None:
        materialized_cells = 0
        merged_expansion = 0
        try:
            worksheet_parts = self._worksheet_parts(archive, infos)
            for info in infos:
                if self._archive_name_key(info.filename) not in worksheet_parts:
                    continue
                with archive.open(info) as stream:
                    saw_root = False
                    for event, element in ElementTree.iterparse(
                        stream,
                        events=("start", "end"),
                    ):
                        tag = element.tag.rsplit("}", 1)[-1]
                        if not saw_root and event == "start":
                            saw_root = True
                            if tag != "worksheet":
                                raise WorkbookParseError(
                                    f"Worksheet part {info.filename!r} has an invalid root element"
                                )
                        if event != "end":
                            continue
                        if tag == "c":
                            reference = element.attrib.get("r", "")
                            min_column, min_row, max_column, max_row = range_boundaries(reference)
                            if (
                                None in {min_column, min_row, max_column, max_row}
                                or int(min_column) < 1
                                or int(max_column) > 16_384
                                or int(min_row) < 1
                                or int(max_row) > 1_048_576
                            ):
                                raise WorkbookParseError(
                                    f"Workbook contains an out-of-bounds cell: {reference!r}"
                                )
                            materialized_cells += 1
                        elif tag == "mergeCell":
                            reference = element.attrib.get("ref", "")
                            min_column, min_row, max_column, max_row = range_boundaries(reference)
                            if None in {min_column, min_row, max_column, max_row}:
                                raise WorkbookParseError(
                                    f"Workbook contains an invalid merged range: {reference!r}"
                                )
                            if (
                                int(min_column) < 1
                                or int(max_column) > 16_384
                                or int(min_row) < 1
                                or int(max_row) > 1_048_576
                            ):
                                raise WorkbookParseError(
                                    f"Merged range {reference!r} falls outside Excel limits"
                                )
                            area = (int(max_column) - int(min_column) + 1) * (
                                int(max_row) - int(min_row) + 1
                            )
                            if area > self.max_merged_cells:
                                raise WorkbookParseError(
                                    f"Merged range {reference!r} expands to {area:,} cells; "
                                    f"the configured limit is {self.max_merged_cells:,}"
                                )
                            merged_expansion += max(0, area - 1)
                        element.clear()
                        if materialized_cells + merged_expansion > self.max_cells:
                            raise WorkbookParseError(
                                "Workbook contains more than the configured "
                                f"{self.max_cells:,} cells after merged-range expansion"
                            )
        except WorkbookParseError:
            raise
        except (ElementTree.ParseError, TypeError, ValueError) as exc:
            raise WorkbookParseError(f"Unable to preflight worksheet XML safely: {exc}") from exc

    def _worksheet_parts(
        self,
        archive: zipfile.ZipFile,
        infos: list[zipfile.ZipInfo],
    ) -> set[str]:
        """Resolve worksheet parts by OPC metadata, independent of filename extension."""

        archive_parts = self._regular_archive_parts(infos)
        parts: set[str] = set()
        defaults, overrides = self._content_types(archive, archive_parts)
        for part_key in archive_parts:
            content_type = self._effective_content_type(part_key, defaults, overrides)
            if content_type and content_type.endswith(_WORKSHEET_CONTENT_TYPE):
                parts.add(part_key)

        for info in infos:
            relationship_name = self._normalise_archive_name(info.filename)
            if not self._is_relationship_part_name(relationship_name):
                continue
            source_part = self._relationship_source_part(relationship_name)
            if source_part is None:
                continue
            for relationship in self._relationships(
                archive,
                info,
                f"relationships part {relationship_name!r}",
            ):
                relationship_type = relationship.attrib["Type"].casefold()
                if not relationship_type.endswith(_WORKSHEET_RELATIONSHIP):
                    continue
                if relationship.attrib.get("TargetMode", "").casefold() == "external":
                    continue
                parts.add(
                    self._resolve_internal_relationship_target(
                        source_part,
                        relationship,
                        archive_parts,
                        "worksheet relationship",
                    )
                )

        return parts

    @staticmethod
    def _relationship_source_part(relationship_name: str) -> str | None:
        folded_name = relationship_name.casefold()
        if folded_name == "_rels/.rels":
            return ""
        marker = "/_rels/"
        marker_index = folded_name.rfind(marker)
        if marker_index < 0:
            return None
        prefix = relationship_name[:marker_index]
        filename = relationship_name[marker_index + len(marker) :]
        if not filename.casefold().endswith(".rels"):
            return None
        return posixpath.join(prefix, filename[: -len(".rels")])

    @staticmethod
    def _normalise_archive_name(name: str) -> str:
        normalised = posixpath.normpath(name.replace("\\", "/").lstrip("/"))
        if normalised == "." or normalised.startswith("../"):
            return ""
        return normalised

    @staticmethod
    def _archive_name_key(name: str) -> str:
        """Return the OPC case-insensitive comparison key for a package part."""

        return WorkbookParser._normalise_archive_name(name).casefold()

    @staticmethod
    def _hidden_rows(dimensions: Iterable[tuple[int, Any]]) -> list[int]:
        hidden: list[int] = []
        for row_index, dimension in dimensions:
            if not bool(dimension.hidden):
                continue
            if row_index < 1 or row_index > 1_048_576:
                raise WorkbookParseError(
                    f"Hidden row {row_index} is outside Excel worksheet limits"
                )
            hidden.append(row_index)
        return sorted(hidden)

    @staticmethod
    def _hidden_columns(dimensions: Iterable[Any]) -> list[str]:
        hidden: set[str] = set()
        for dimension in dimensions:
            if not bool(dimension.hidden):
                continue
            minimum = int(dimension.min or 0)
            maximum = int(dimension.max or minimum)
            if minimum < 1 or maximum < minimum or maximum > 16_384:
                raise WorkbookParseError(
                    f"Hidden column range {minimum}:{maximum} is outside Excel limits"
                )
            hidden.update(get_column_letter(index) for index in range(minimum, maximum + 1))
        return sorted(hidden)

    @staticmethod
    def _data_validations(worksheet: Any) -> list[DataValidationInfo]:
        container = getattr(worksheet, "data_validations", None)
        validations = getattr(container, "dataValidation", []) if container else []
        result = [
            DataValidationInfo(
                ranges=str(item.sqref),
                validation_type=item.type,
                operator=item.operator,
                formula1=None if item.formula1 is None else str(item.formula1),
                formula2=None if item.formula2 is None else str(item.formula2),
                allow_blank=bool(item.allowBlank),
                show_error_message=bool(item.showErrorMessage),
            )
            for item in validations
        ]
        return sorted(result, key=lambda item: (item.ranges, item.validation_type or ""))

    @staticmethod
    def _named_ranges(workbook: Any) -> list[NamedRangeInfo]:
        values = getattr(workbook.defined_names, "values", None)
        definitions = values() if callable(values) else []
        result = [
            NamedRangeInfo(
                name=str(item.name),
                value=str(item.attr_text or ""),
                kind=getattr(item, "type", None),
                local_sheet_id=getattr(item, "localSheetId", None),
                hidden=getattr(item, "hidden", None),
            )
            for item in definitions
        ]
        return sorted(result, key=lambda item: (item.name, item.local_sheet_id or -1))

    @staticmethod
    def _external_links(workbook: Any, sheets: list[SheetSnapshot]) -> list[str]:
        links: set[str] = set()
        for link in getattr(workbook, "_external_links", []):
            file_link = getattr(link, "file_link", None)
            target = getattr(file_link, "Target", None)
            links.add(str(target or link))
        for sheet in sheets:
            for cell in sheet.cells.values():
                if cell.formula:
                    links.update(_EXTERNAL_BOOK_RE.findall(cell.formula))
        return sorted(links)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
