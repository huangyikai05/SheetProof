"""Deterministic implementations of Tabulint's built-in rules."""

from __future__ import annotations

import math
from dataclasses import dataclass
from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from tabulint.models import (
    CellChange,
    CellKind,
    RuleResult,
    RuleSpec,
    RuleStatus,
    RuleType,
    SheetSnapshot,
    StructureChange,
    WorkbookSnapshot,
)

_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COLUMN = 16_384
_EVIDENCE_SAMPLE_LIMIT = 100


class RuleEvaluationError(ValueError):
    """Raised when a rule cannot be evaluated reliably."""


@dataclass(frozen=True, slots=True)
class RuleContext:
    """Immutable facts available to every built-in rule."""

    before: WorkbookSnapshot
    after: WorkbookSnapshot
    structure_changes: list[StructureChange]
    cell_changes: list[CellChange]
    max_range_cells: int


@dataclass(frozen=True, slots=True)
class ParsedRange:
    """A validated, bounded A1 range with an explicit worksheet."""

    raw: str
    sheet: str
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def cell_count(self) -> int:
        return (self.max_column - self.min_column + 1) * (self.max_row - self.min_row + 1)

    def contains(self, sheet: str, coordinate: str) -> bool:
        if sheet.casefold() != self.sheet.casefold():
            return False
        min_column, min_row, max_column, max_row = _coordinate_bounds(coordinate)
        return (
            min_column == max_column
            and min_row == max_row
            and self.min_column <= min_column <= self.max_column
            and self.min_row <= min_row <= self.max_row
        )


def evaluate_builtin_rule(spec: RuleSpec, context: RuleContext) -> RuleResult:
    """Dispatch one validated rule specification to its implementation."""

    evaluators = {
        RuleType.FORMULA_REQUIRED: _formula_required,
        RuleType.ALLOWED_CHANGE_RANGE: _allowed_change_range,
        RuleType.NO_EXTERNAL_LINKS: _no_external_links,
        RuleType.NO_NEW_HIDDEN_SHEETS: _no_new_hidden_sheets,
        RuleType.NO_MACRO_ADDED: _no_macro_added,
        RuleType.NUMERIC_RANGE: _numeric_range,
        RuleType.REQUIRED_SHEET: _required_sheet,
        RuleType.FORBIDDEN_SHEET: _forbidden_sheet,
        RuleType.MAX_CHANGED_CELLS: _max_changed_cells,
    }
    evaluator = evaluators.get(spec.type)
    if evaluator is None:
        raise RuleEvaluationError(f"Unsupported rule type: {spec.type}")
    return evaluator(spec, context)


def error_result(spec: RuleSpec, reason: str, *, error_type: str) -> RuleResult:
    """Create a visible result for a rule that could not be evaluated."""

    return RuleResult(
        name=spec.name,
        rule_type=spec.type.value,
        status=RuleStatus.ERROR,
        severity=spec.severity,
        reason=reason,
        location=_rule_location(spec),
        evidence={"error_type": error_type},
    )


def _formula_required(spec: RuleSpec, context: RuleContext) -> RuleResult:
    assert spec.range is not None
    target = parse_range(spec.range)
    sheet = _sheet(context.after, target.sheet)
    if sheet is None:
        raise RuleEvaluationError(
            f"Worksheet '{target.sheet}' does not exist in the reviewed workbook"
        )
    if target.cell_count > context.max_range_cells:
        raise RuleEvaluationError(
            f"Range '{target.raw}' contains {target.cell_count:,} cells; "
            f"the evaluation limit is {context.max_range_cells:,}"
        )

    non_formula_cells: list[str] = []
    for row in range(target.min_row, target.max_row + 1):
        for column in range(target.min_column, target.max_column + 1):
            coordinate = f"{get_column_letter(column)}{row}"
            cell = sheet.cells.get(coordinate)
            if cell is None or cell.kind is not CellKind.FORMULA or cell.formula is None:
                non_formula_cells.append(coordinate)

    count = len(non_formula_cells)
    evidence: dict[str, Any] = {
        "range": target.raw,
        "checked_cells": target.cell_count,
        "non_formula_count": count,
        "non_formula_cells": non_formula_cells[:_EVIDENCE_SAMPLE_LIMIT],
        "evidence_truncated": count > _EVIDENCE_SAMPLE_LIMIT,
    }
    if count:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"{count} of {target.cell_count} required cells do not contain formulas.",
            location=target.raw,
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"All {target.cell_count} required cells contain formulas.",
        location=target.raw,
        evidence=evidence,
    )


def _allowed_change_range(spec: RuleSpec, context: RuleContext) -> RuleResult:
    allowed = [parse_range(value) for value in spec.ranges]
    available_sheets = {
        sheet.name.casefold() for sheet in [*context.before.sheets, *context.after.sheets]
    }
    missing_sheets = sorted(
        {item.sheet for item in allowed if item.sheet.casefold() not in available_sheets}
    )
    if missing_sheets:
        names = ", ".join(repr(name) for name in missing_sheets)
        raise RuleEvaluationError(f"Allowed range refers to missing worksheet(s): {names}")

    unique_changes = _unique_cell_changes(context.cell_changes)
    violations: list[str] = []
    for change in unique_changes:
        if not any(item.contains(change.sheet, change.coordinate) for item in allowed):
            violations.append(change.location)

    evidence: dict[str, Any] = {
        "scope": "cell_changes",
        "allowed_ranges": spec.ranges,
        "changed_cell_count": len(unique_changes),
        "outside_allowed_count": len(violations),
        "outside_allowed_cells": violations[:_EVIDENCE_SAMPLE_LIMIT],
        "evidence_truncated": len(violations) > _EVIDENCE_SAMPLE_LIMIT,
    }
    location = ", ".join(spec.ranges)
    if violations:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"{len(violations)} changed cells are outside the allowed ranges.",
            location=location,
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"All {len(unique_changes)} changed cells are within the allowed ranges.",
        location=location,
        evidence=evidence,
    )


def _no_external_links(spec: RuleSpec, context: RuleContext) -> RuleResult:
    before_links = set(context.before.external_links)
    after_links = set(context.after.external_links)
    additions = sorted(after_links - before_links)
    evidence: dict[str, Any] = {
        "before_links": sorted(before_links),
        "after_links": sorted(after_links),
        "added_links": additions,
        "matching_structure_changes": _matching_structure_locations(
            context, "external_link_added"
        ),
    }
    if additions:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"{len(additions)} external link(s) were added.",
            location="Workbook",
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        "No external links were added.",
        location="Workbook",
        evidence=evidence,
    )


def _no_new_hidden_sheets(spec: RuleSpec, context: RuleContext) -> RuleResult:
    before_names = {sheet.name.casefold() for sheet in context.before.sheets}
    added_hidden = sorted(
        sheet.name
        for sheet in context.after.sheets
        if sheet.name.casefold() not in before_names and sheet.state != "visible"
    )
    before_hidden = sorted(
        sheet.name for sheet in context.before.sheets if sheet.state != "visible"
    )
    after_hidden = sorted(sheet.name for sheet in context.after.sheets if sheet.state != "visible")
    evidence: dict[str, Any] = {
        "before_hidden_sheets": before_hidden,
        "after_hidden_sheets": after_hidden,
        "added_hidden_sheets": added_hidden,
        "matching_structure_changes": _matching_structure_locations(
            context, "hidden_sheet_added"
        ),
    }
    if added_hidden:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"{len(added_hidden)} hidden sheet(s) were added.",
            location="Workbook",
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        "No hidden sheets were added.",
        location="Workbook",
        evidence=evidence,
    )


def _no_macro_added(spec: RuleSpec, context: RuleContext) -> RuleResult:
    added = not context.before.has_vba and context.after.has_vba
    evidence: dict[str, Any] = {
        "before_has_vba": context.before.has_vba,
        "after_has_vba": context.after.has_vba,
        "macro_added": added,
        "matching_structure_changes": _matching_structure_locations(context, "macro_added"),
    }
    if added:
        return _result(
            spec,
            RuleStatus.FAILED,
            "A VBA macro project was added to the workbook.",
            location="Workbook",
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        "No VBA macro project was added.",
        location="Workbook",
        evidence=evidence,
    )


def _numeric_range(spec: RuleSpec, context: RuleContext) -> RuleResult:
    assert spec.target is not None
    target = parse_range(spec.target, require_single_cell=True)
    if spec.min is not None and not math.isfinite(spec.min):
        raise RuleEvaluationError("numeric_range minimum must be finite")
    if spec.max is not None and not math.isfinite(spec.max):
        raise RuleEvaluationError("numeric_range maximum must be finite")
    if spec.min is not None and spec.max is not None and spec.min > spec.max:
        raise RuleEvaluationError("numeric_range minimum cannot exceed maximum")

    sheet = _sheet(context.after, target.sheet)
    if sheet is None:
        raise RuleEvaluationError(
            f"Worksheet '{target.sheet}' does not exist in the reviewed workbook"
        )
    coordinate = f"{get_column_letter(target.min_column)}{target.min_row}"
    cell = sheet.cells.get(coordinate)
    if cell is None:
        raise RuleEvaluationError(
            f"Target cell '{target.raw}' does not exist in the parsed workbook"
        )

    value = cell.cached_value if cell.kind is CellKind.FORMULA else cell.value
    source = "cached_formula_value" if cell.kind is CellKind.FORMULA else "cell_value"
    evidence: dict[str, Any] = {
        "target": target.raw,
        "cell_kind": cell.kind.value,
        "value_source": source,
        "value": value,
        "minimum": spec.min,
        "maximum": spec.max,
    }
    if cell.kind is CellKind.FORMULA and value is None:
        return _result(
            spec,
            RuleStatus.SKIPPED,
            "The formula has no cached value; this version cannot reliably recalculate it.",
            location=target.raw,
            evidence=evidence,
        )
    if isinstance(value, bool) or not isinstance(value, int | float):
        return _result(
            spec,
            RuleStatus.FAILED,
            "The target does not contain a numeric value that can be checked.",
            location=target.raw,
            evidence=evidence,
        )
    if not math.isfinite(float(value)):
        return _result(
            spec,
            RuleStatus.ERROR,
            "The target numeric value is not finite and cannot be checked reliably.",
            location=target.raw,
            evidence={**evidence, "error_type": "non_finite_value"},
        )

    below_minimum = spec.min is not None and value < spec.min
    above_maximum = spec.max is not None and value > spec.max
    checked_evidence: dict[str, Any] = {
        **evidence,
        "below_minimum": below_minimum,
        "above_maximum": above_maximum,
    }
    if below_minimum or above_maximum:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"Numeric value {value!r} is outside the configured inclusive range.",
            location=target.raw,
            evidence=checked_evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"Numeric value {value!r} is within the configured inclusive range.",
        location=target.raw,
        evidence=checked_evidence,
    )


def _required_sheet(spec: RuleSpec, context: RuleContext) -> RuleResult:
    assert spec.sheet is not None
    actual_names = [sheet.name for sheet in context.after.sheets]
    found = any(name.casefold() == spec.sheet.casefold() for name in actual_names)
    evidence: dict[str, Any] = {
        "required_sheet": spec.sheet,
        "available_sheets": actual_names,
        "found": found,
    }
    if not found:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"Required worksheet '{spec.sheet}' does not exist.",
            location=spec.sheet,
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"Required worksheet '{spec.sheet}' exists.",
        location=spec.sheet,
        evidence=evidence,
    )


def _forbidden_sheet(spec: RuleSpec, context: RuleContext) -> RuleResult:
    assert spec.sheet is not None
    actual_names = [sheet.name for sheet in context.after.sheets]
    found = any(name.casefold() == spec.sheet.casefold() for name in actual_names)
    evidence: dict[str, Any] = {
        "forbidden_sheet": spec.sheet,
        "available_sheets": actual_names,
        "found": found,
    }
    if found:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"Forbidden worksheet '{spec.sheet}' exists.",
            location=spec.sheet,
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"Forbidden worksheet '{spec.sheet}' is absent.",
        location=spec.sheet,
        evidence=evidence,
    )


def _max_changed_cells(spec: RuleSpec, context: RuleContext) -> RuleResult:
    assert spec.max is not None
    maximum = int(spec.max)
    changes = _unique_cell_changes(context.cell_changes)
    count = len(changes)
    sample = [change.location for change in changes[:_EVIDENCE_SAMPLE_LIMIT]]
    evidence: dict[str, Any] = {
        "maximum_changed_cells": maximum,
        "changed_cell_count": count,
        "changed_cells": sample,
        "evidence_truncated": count > _EVIDENCE_SAMPLE_LIMIT,
    }
    if count > maximum:
        return _result(
            spec,
            RuleStatus.FAILED,
            f"{count} cells changed, exceeding the configured maximum of {maximum}.",
            location="Workbook",
            evidence=evidence,
        )
    return _result(
        spec,
        RuleStatus.PASSED,
        f"{count} cells changed, within the configured maximum of {maximum}.",
        location="Workbook",
        evidence=evidence,
    )


def parse_range(value: str, *, require_single_cell: bool = False) -> ParsedRange:
    """Parse ``Sheet!A1:B2`` without resolving names or touching a workbook."""

    raw = value.strip()
    if "!" not in raw:
        raise RuleEvaluationError(f"Range '{value}' must include a worksheet name (Sheet!A1)")
    sheet_text, cell_text = raw.rsplit("!", 1)
    sheet = _unquote_sheet_name(sheet_text.strip())
    if not sheet:
        raise RuleEvaluationError(f"Range '{value}' has an empty worksheet name")
    try:
        min_column, min_row, max_column, max_row = range_boundaries(cell_text.replace("$", ""))
    except (TypeError, ValueError) as exc:
        raise RuleEvaluationError(f"Range '{value}' is not valid A1 notation") from exc
    bounds = (min_column, min_row, max_column, max_row)
    if any(item is None for item in bounds):
        raise RuleEvaluationError(f"Range '{value}' must have bounded rows and columns")
    if not (
        1 <= min_column <= max_column <= _EXCEL_MAX_COLUMN
        and 1 <= min_row <= max_row <= _EXCEL_MAX_ROW
    ):
        raise RuleEvaluationError(f"Range '{value}' falls outside Excel worksheet limits")
    if require_single_cell and (min_column != max_column or min_row != max_row):
        raise RuleEvaluationError(f"Target '{value}' must identify exactly one cell")
    return ParsedRange(raw, sheet, min_column, min_row, max_column, max_row)


def _coordinate_bounds(coordinate: str) -> tuple[int, int, int, int]:
    try:
        bounds = range_boundaries(coordinate.replace("$", ""))
    except (TypeError, ValueError) as exc:
        raise RuleEvaluationError(f"Invalid changed-cell coordinate: {coordinate!r}") from exc
    if any(item is None for item in bounds):
        raise RuleEvaluationError(f"Invalid changed-cell coordinate: {coordinate!r}")
    min_column, min_row, max_column, max_row = bounds
    return int(min_column), int(min_row), int(max_column), int(max_row)


def _unquote_sheet_name(value: str) -> str:
    if value.startswith("'") and value.endswith("'") and len(value) >= 2:
        return value[1:-1].replace("''", "'")
    return value


def _sheet(workbook: WorkbookSnapshot, name: str) -> SheetSnapshot | None:
    wanted = name.casefold()
    return next((sheet for sheet in workbook.sheets if sheet.name.casefold() == wanted), None)


def _unique_cell_changes(changes: list[CellChange]) -> list[CellChange]:
    unique: dict[tuple[str, str], CellChange] = {}
    for change in changes:
        key = (change.sheet.casefold(), change.coordinate.replace("$", "").upper())
        unique.setdefault(key, change)
    return list(unique.values())


def _matching_structure_locations(context: RuleContext, change_type: str) -> list[str]:
    return sorted(
        {
            change.location
            for change in context.structure_changes
            if change.change_type == change_type
        }
    )


def _rule_location(spec: RuleSpec) -> str | None:
    if spec.range:
        return spec.range
    if spec.ranges:
        return ", ".join(spec.ranges)
    if spec.target:
        return spec.target
    if spec.sheet:
        return spec.sheet
    return "Workbook"


def _result(
    spec: RuleSpec,
    status: RuleStatus,
    reason: str,
    *,
    location: str | None,
    evidence: dict[str, Any],
) -> RuleResult:
    return RuleResult(
        name=spec.name,
        rule_type=spec.type.value,
        status=status,
        severity=spec.severity,
        reason=reason,
        location=location,
        evidence=evidence,
    )


__all__ = [
    "ParsedRange",
    "RuleContext",
    "RuleEvaluationError",
    "error_result",
    "evaluate_builtin_rule",
    "parse_range",
]
