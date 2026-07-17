"""Generate safe, deterministic Tabulint demo workbooks from source."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_DIR = Path(__file__).resolve().parent / "generated"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def _style_header(sheet: object, cell_range: str) -> None:
    for row in sheet[cell_range]:  # type: ignore[index]
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center")


def _build_before() -> Workbook:
    workbook = Workbook()
    forecast = workbook.active
    forecast.title = "预测"
    cashflow = workbook.create_sheet("现金流")
    profit = workbook.create_sheet("利润表")
    notes = workbook.create_sheet("说明")

    forecast.merge_cells("A1:H1")
    forecast["A1"] = "Tabulint Forecast Demo"
    forecast["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    forecast["A1"].fill = HEADER_FILL
    forecast["A1"].alignment = Alignment(horizontal="center")
    forecast.append([])
    forecast["A3"] = "Yellow cells are the approved forecast input area."
    headers = ["Period", "Department", "Scenario", "Revenue", "Cost", "Tax", "Capex", "Cash"]
    for column, value in enumerate(headers, start=1):
        forecast.cell(row=4, column=column, value=value)
    _style_header(forecast, "A4:H4")
    departments = ["North", "South", "Online"]
    for row in range(5, 31):
        forecast.cell(row=row, column=1, value=f"M{row - 4:02d}")
        forecast.cell(row=row, column=2, value=departments[(row - 5) % len(departments)])
        forecast.cell(row=row, column=3, value="Base")
        for column in range(4, 9):
            value = (row - 4) * 1000 + column * 25
            cell = forecast.cell(row=row, column=column, value=value)
            cell.fill = INPUT_FILL
            cell.number_format = '#,##0'
    validation = DataValidation(type="list", formula1='"Base,Upside,Downside"', allow_blank=False)
    forecast.add_data_validation(validation)
    validation.add("C5:C30")
    forecast.conditional_formatting.add(
        "D5:H30",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    forecast.freeze_panes = "D5"
    forecast.column_dimensions["A"].width = 12
    forecast.column_dimensions["B"].width = 14
    forecast.column_dimensions["C"].width = 12
    for column in "DEFGH":
        forecast.column_dimensions[column].width = 13
    table = Table(displayName="ForecastInputs", ref="A4:H30")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    forecast.add_table(table)

    cashflow["A1"] = "Cash Flow Review Area"
    cashflow["A1"].font = Font(size=15, bold=True)
    cashflow["A4"], cashflow["B4"] = "Period", "Net cash flow"
    _style_header(cashflow, "A4:B4")
    for row in range(5, 21):
        cashflow.cell(row=row, column=1, value=f"M{row - 4:02d}")
        cashflow.cell(
            row=row,
            column=2,
            value=f"='预测'!D{row}-'预测'!E{row}-'预测'!F{row}-'预测'!G{row}",
        ).number_format = '#,##0'
    cashflow["A22"], cashflow["B22"] = "Total", "=SUM(B5:B20)"
    cashflow["A22"].fill = SECTION_FILL
    cashflow["B22"].fill = SECTION_FILL
    cashflow["B22"].font = Font(bold=True)
    cashflow.freeze_panes = "B5"
    cashflow.column_dimensions["A"].width = 14
    cashflow.column_dimensions["B"].width = 18

    profit["A1"] = "Profit and Loss"
    profit["A1"].font = Font(size=15, bold=True)
    profit["E10"] = 100000
    profit["E10"].number_format = '#,##0'
    profit["E18"] = 28000
    profit["E18"].number_format = '#,##0'
    profit["F10"] = "Revenue"
    profit["F17"] = "Net profit"
    profit["F18"] = "=E18/E10"
    profit["F18"].number_format = "0.0%"
    profit.column_dimensions["E"].width = 14
    profit.column_dimensions["F"].width = 18

    notes["A1"] = "Reviewer notes"
    notes["A1"].font = Font(size=15, bold=True)
    notes["A2"] = "Owner"
    notes["B2"] = "Finance Operations"
    notes["A3"] = "Purpose"
    notes["B3"] = "Demonstrate deterministic workbook change review."
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 54

    workbook.defined_names.add(DefinedName("CoreMargin", attr_text="'利润表'!$F$18"))
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def generate(output_dir: Path = OUTPUT_DIR) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    before_path = output_dir / "before.xlsx"
    safe_path = output_dir / "after_safe.xlsx"
    risky_path = output_dir / "after_risky.xlsx"

    before = _build_before()
    before.save(before_path)

    safe = load_workbook(before_path)
    safe["预测"]["D5"] = 1125
    safe.save(safe_path)
    safe.close()

    risky = load_workbook(before_path)
    risky["现金流"]["B13"] = 12500
    risky["现金流"]["B22"] = "=SUM(B5:B19)"
    risky["说明"]["B2"] = "Unapproved automation"
    hidden = risky.create_sheet("隐藏调整")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "='[external.xlsx]Inputs'!A1"
    hidden["A2"] = "This sheet and link were intentionally added for review."
    risky.save(risky_path)
    risky.close()

    paths = [before_path, safe_path, risky_path]
    for path in paths:
        print(path)
    return paths


if __name__ == "__main__":
    generate()
