"""Read workbook facts without executing macros, formulas, or external content."""

from __future__ import annotations

import hashlib
import json
import math
import posixpath
import re
import struct
import zipfile
from collections.abc import Iterable
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from sheetproof.exceptions import WorkbookParseError
from sheetproof.models import (
    CellKind,
    CellSnapshot,
    DataValidationInfo,
    FileInfo,
    FormulaCalculationStatus,
    NamedRangeInfo,
    ScalarValue,
    SheetSnapshot,
    StyleSummary,
    TableInfo,
    WorkbookSnapshot,
)

_EXTERNAL_BOOK_RE = re.compile(r"\[([^\]]+\.(?:xlsx|xlsm|xlsb|xls))\]", re.IGNORECASE)
_SUPPORTED_SUFFIXES = {".xlsx", ".xlsm"}
_WORKSHEET_RELATIONSHIP = "/worksheet"
_WORKSHEET_CONTENT_TYPE = "/worksheet+xml"
_EOCD_SIGNATURE = b"PK\x05\x06"
_CENTRAL_FILE_SIGNATURE = b"PK\x01\x02"
_CENTRAL_SIGNATURE_SIGNATURE = b"PK\x05\x05"
_MAX_ZIP_COMMENT = 65_535


class WorkbookParser:
    """Parse deterministic workbook metadata with explicit resource limits."""

    def __init__(
        self,
        *,
        max_file_size: int = 100 * 1024 * 1024,
        max_uncompressed_size: int = 512 * 1024 * 1024,
        max_archive_entries: int = 10_000,
        max_cells: int = 1_000_000,
        max_merged_cells: int = 100_000,
    ) -> None:
        limits = (
            max_file_size,
            max_uncompressed_size,
            max_archive_entries,
            max_cells,
            max_merged_cells,
        )
        if any(
            isinstance(value, bool) or not isinstance(value, int) or value < 1
            for value in limits
        ):
            raise ValueError("Workbook parser limits must be positive integers")
        self.max_file_size = max_file_size
        self.max_uncompressed_size = max_uncompressed_size
        self.max_archive_entries = max_archive_entries
        self.max_cells = max_cells
        self.max_merged_cells = max_merged_cells

    def parse(self, workbook_path: str | Path) -> WorkbookSnapshot:
        path = Path(workbook_path).expanduser().resolve()
        self._validate_path(path)
        archive_names = self._validate_archive(path)
        has_vba = "xl/vbaProject.bin" in archive_names

        try:
            workbook = load_workbook(
                path,
                data_only=False,
                read_only=False,
                keep_vba=has_vba,
                keep_links=True,
            )
            cached_workbook = load_workbook(
                path,
                data_only=True,
                read_only=False,
                keep_vba=False,
                keep_links=False,
            )
        except Exception as exc:
            raise WorkbookParseError(f"Unable to open workbook '{path.name}': {exc}") from exc

        try:
            sheets: list[SheetSnapshot] = []
            parsed_cells = 0
            for index, worksheet in enumerate(workbook.worksheets):
                cached_sheet = cached_workbook[worksheet.title]
                cells: dict[str, CellSnapshot] = {}
                raw_cells: Iterable[Any] = getattr(worksheet, "_cells", {}).values()
                for cell in raw_cells:
                    if isinstance(cell, MergedCell):
                        continue
                    parsed_cells += 1
                    if parsed_cells > self.max_cells:
                        raise WorkbookParseError(
                            f"Workbook contains more than the configured {self.max_cells:,} cells"
                        )
                    cached_value = None
                    if self._is_formula(cell.value, cell.data_type):
                        cached_value = self._normalise_value(cached_sheet[cell.coordinate].value)
                    cells[cell.coordinate] = self._cell_snapshot(cell, cached_value)

                sheets.append(
                    SheetSnapshot(
                        name=worksheet.title,
                        index=index,
                        state=worksheet.sheet_state,
                        cells=cells,
                        hidden_rows=self._hidden_rows(worksheet.row_dimensions.items()),
                        hidden_columns=self._hidden_columns(worksheet.column_dimensions.values()),
                        merged_cells=sorted(str(item) for item in worksheet.merged_cells.ranges),
                        data_validations=self._data_validations(worksheet),
                        freeze_panes=(
                            str(worksheet.freeze_panes)
                            if worksheet.freeze_panes is not None
                            else None
                        ),
                        tables=sorted(
                            (
                                TableInfo(
                                    name=table.name,
                                    display_name=table.displayName,
                                    ref=table.ref,
                                    totals_row_shown=table.totalsRowShown,
                                )
                                for table in worksheet.tables.values()
                            ),
                            key=lambda item: item.name,
                        ),
                    )
                )

            external_links = self._external_links(workbook, sheets)
            digest = self._sha256(path)
            return WorkbookSnapshot(
                file=FileInfo.from_path(path, sha256=digest, has_vba=has_vba),
                sheets=sheets,
                named_ranges=self._named_ranges(workbook),
                external_links=external_links,
                has_vba=has_vba,
                parsed_cell_count=parsed_cells,
            )
        finally:
            workbook.close()
            cached_workbook.close()

    def _validate_path(self, path: Path) -> None:
        if not path.exists():
            raise WorkbookParseError(f"Workbook does not exist: {path}")
        if not path.is_file():
            raise WorkbookParseError(f"Workbook path is not a file: {path}")
        if path.suffix.lower() not in _SUPPORTED_SUFFIXES:
            raise WorkbookParseError("Only .xlsx and .xlsm workbooks are supported")
        size = path.stat().st_size
        if size > self.max_file_size:
            raise WorkbookParseError(
                f"Workbook is {size:,} bytes; maximum allowed size is {self.max_file_size:,}"
            )

    def _validate_archive(self, path: Path) -> set[str]:
        self._preflight_central_directory(path)
        try:
            with zipfile.ZipFile(path) as archive:
                infos = archive.infolist()
                if len(infos) > self.max_archive_entries:
                    raise WorkbookParseError(
                        "Workbook archive contains more than the configured "
                        f"{self.max_archive_entries:,} entries"
                    )
                normalised_names: set[str] = set()
                for info in infos:
                    raw_name = info.filename
                    normalised_name = self._normalise_archive_name(raw_name)
                    if (
                        not normalised_name
                        or raw_name.startswith(("/", "\\"))
                        or "\\" in raw_name
                        or "//" in raw_name
                        or any(part in {".", ".."} for part in raw_name.split("/"))
                        or normalised_name in normalised_names
                    ):
                        raise WorkbookParseError(
                            f"Workbook contains an unsafe or duplicate archive path: {raw_name!r}"
                        )
                    normalised_names.add(normalised_name)
                total_uncompressed = sum(info.file_size for info in infos)
                if total_uncompressed > self.max_uncompressed_size:
                    raise WorkbookParseError(
                        "Workbook archive expands beyond the configured safety limit"
                    )
                for info in infos:
                    if info.compress_size and info.file_size / info.compress_size > 2_000:
                        raise WorkbookParseError(
                            "Workbook contains a suspiciously compressed entry"
                        )
                self._preflight_sheet_xml(archive, infos)
                return {info.filename for info in infos}
        except WorkbookParseError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise WorkbookParseError(f"Workbook is not a valid OOXML archive: {exc}") from exc

    def _preflight_central_directory(self, path: Path) -> None:
        """Stream central headers before ``ZipFile`` can allocate a large entry list."""

        try:
            file_size = path.stat().st_size
            tail_size = min(file_size, _MAX_ZIP_COMMENT + 22)
            with path.open("rb") as stream:
                stream.seek(file_size - tail_size)
                tail = stream.read(tail_size)
                eocd_offset = self._find_eocd(tail)
                if eocd_offset is None:
                    raise WorkbookParseError(
                        "Workbook is not a valid OOXML archive: missing ZIP end record"
                    )

                (
                    disk_number,
                    directory_disk,
                    disk_entries,
                    stated_entries,
                    directory_size,
                    _directory_offset,
                ) = struct.unpack_from("<4H2L", tail, eocd_offset + 4)
                if disk_number or directory_disk or disk_entries != stated_entries:
                    raise WorkbookParseError("Multi-disk workbook archives are not supported")
                if (
                    stated_entries == 0xFFFF
                    or directory_size == 0xFFFFFFFF
                    or _directory_offset == 0xFFFFFFFF
                ):
                    raise WorkbookParseError("ZIP64 workbook archives are not supported")

                absolute_eocd = file_size - tail_size + eocd_offset
                directory_start = absolute_eocd - directory_size
                if directory_start < 0:
                    raise WorkbookParseError("Workbook central directory has invalid bounds")
                stream.seek(directory_start)
                remaining = directory_size
                counted_entries = 0
                while remaining:
                    if remaining < 4:
                        raise WorkbookParseError("Workbook central directory is truncated")
                    signature = stream.read(4)
                    remaining -= 4
                    if signature == _CENTRAL_FILE_SIGNATURE:
                        if remaining < 42:
                            raise WorkbookParseError(
                                "Workbook central directory header is truncated"
                            )
                        fixed = stream.read(42)
                        remaining -= 42
                        name_size, extra_size, comment_size = struct.unpack_from(
                            "<3H", fixed, 24
                        )
                        variable_size = name_size + extra_size + comment_size
                        if variable_size > remaining:
                            raise WorkbookParseError(
                                "Workbook central directory entry has invalid bounds"
                            )
                        stream.seek(variable_size, 1)
                        remaining -= variable_size
                        counted_entries += 1
                        if counted_entries > self.max_archive_entries:
                            raise WorkbookParseError(
                                "Workbook archive contains more than the configured "
                                f"{self.max_archive_entries:,} entries"
                            )
                    elif signature == _CENTRAL_SIGNATURE_SIGNATURE:
                        if remaining < 2:
                            raise WorkbookParseError(
                                "Workbook central-directory signature is truncated"
                            )
                        signature_size = struct.unpack("<H", stream.read(2))[0]
                        remaining -= 2
                        if signature_size != remaining:
                            raise WorkbookParseError(
                                "Workbook central-directory signature has invalid bounds"
                            )
                        stream.seek(signature_size, 1)
                        remaining = 0
                    else:
                        raise WorkbookParseError(
                            "Workbook central directory contains an invalid record"
                        )
                if counted_entries != stated_entries:
                    raise WorkbookParseError(
                        "Workbook central-directory entry count does not match its end record"
                    )
        except WorkbookParseError:
            raise
        except (OSError, struct.error) as exc:
            raise WorkbookParseError(f"Unable to preflight workbook archive: {exc}") from exc

    @staticmethod
    def _find_eocd(tail: bytes) -> int | None:
        search_end = len(tail)
        while search_end >= 0:
            offset = tail.rfind(_EOCD_SIGNATURE, 0, search_end)
            if offset < 0:
                return None
            if len(tail) >= offset + 22:
                comment_size = struct.unpack_from("<H", tail, offset + 20)[0]
                if offset + 22 + comment_size == len(tail):
                    return offset
            search_end = offset
        return None

    def _cell_snapshot(self, cell: Any, cached_value: ScalarValue) -> CellSnapshot:
        formula, formula_attributes = self._formula_details(cell.value, cell.data_type)
        raw_value = formula if formula is not None else self._normalise_value(cell.value)
        calculation_status = None
        if formula is not None:
            calculation_status = (
                FormulaCalculationStatus.CACHED
                if cached_value is not None
                else FormulaCalculationStatus.UNCALCULATED
            )
        return CellSnapshot(
            coordinate=cell.coordinate,
            value=raw_value,
            formula=formula,
            formula_attributes=formula_attributes,
            cached_value=cached_value,
            calculation_status=calculation_status,
            kind=self._cell_kind(cell, formula),
            data_type=str(cell.data_type),
            is_date=bool(cell.is_date),
            style=self._style_summary(cell),
        )

    @staticmethod
    def _is_formula(value: object, data_type: object) -> bool:
        return data_type == "f" or (isinstance(value, str) and value.startswith("="))

    @staticmethod
    def _formula_details(
        value: object,
        data_type: object,
    ) -> tuple[str | None, dict[str, ScalarValue]]:
        if not WorkbookParser._is_formula(value, data_type):
            return None, {}
        if isinstance(value, ArrayFormula):
            return value.text or "=UNSUPPORTED_ARRAY_FORMULA()", {
                "kind": "array",
                "ref": value.ref,
            }
        if isinstance(value, DataTableFormula):
            return "=UNSUPPORTED_DATA_TABLE()", {
                "kind": "data_table",
                "ref": value.ref,
                "ca": bool(value.ca),
                "dt2D": bool(value.dt2D),
                "dtr": bool(value.dtr),
                "r1": value.r1,
                "r2": value.r2,
                "del1": bool(value.del1),
                "del2": bool(value.del2),
            }
        return str(value), {}

    @staticmethod
    def _cell_kind(cell: Any, formula: str | None) -> CellKind:
        if formula is not None:
            return CellKind.FORMULA
        if cell.value is None:
            return CellKind.BLANK
        if cell.data_type == "e":
            return CellKind.ERROR
        if bool(cell.is_date):
            return CellKind.DATE
        if isinstance(cell.value, bool):
            return CellKind.BOOLEAN
        if isinstance(cell.value, int | float):
            return CellKind.NUMBER
        if isinstance(cell.value, str):
            return CellKind.TEXT
        return CellKind.OTHER

    @staticmethod
    def _normalise_value(value: object) -> ScalarValue:
        if isinstance(value, float) and not math.isfinite(value):
            raise WorkbookParseError("Workbook contains a non-finite numeric cell value")
        if value is None or isinstance(value, str | int | float | bool):
            return value
        if isinstance(value, datetime | date | time):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _style_summary(cell: Any) -> StyleSummary:
        border_sides = {
            side: {
                "style": getattr(getattr(cell.border, side, None), "style", None),
                "color": WorkbookParser._color_summary(
                    getattr(getattr(cell.border, side, None), "color", None)
                ),
            }
            for side in ("left", "right", "top", "bottom", "diagonal")
        }
        return StyleSummary(
            style_id=int(cell.style_id),
            number_format=str(cell.number_format),
            font=json.dumps(
                {
                    "name": cell.font.name,
                    "size": cell.font.sz,
                    "bold": cell.font.b,
                    "italic": cell.font.i,
                    "underline": cell.font.u,
                    "strike": cell.font.strike,
                    "vertical_alignment": cell.font.vertAlign,
                    "color": WorkbookParser._color_summary(cell.font.color),
                },
                sort_keys=True,
                default=str,
            ),
            fill=json.dumps(
                {
                    "type": cell.fill.fill_type,
                    "foreground": WorkbookParser._color_summary(cell.fill.fgColor),
                    "background": WorkbookParser._color_summary(cell.fill.bgColor),
                },
                sort_keys=True,
                default=str,
            ),
            border=json.dumps(
                {
                    "sides": border_sides,
                    "diagonal_up": cell.border.diagonalUp,
                    "diagonal_down": cell.border.diagonalDown,
                    "outline": cell.border.outline,
                },
                sort_keys=True,
                default=str,
            ),
            alignment=json.dumps(
                {
                    "horizontal": cell.alignment.horizontal,
                    "vertical": cell.alignment.vertical,
                    "wrap": cell.alignment.wrap_text,
                    "rotation": cell.alignment.textRotation,
                    "shrink_to_fit": cell.alignment.shrinkToFit,
                    "indent": cell.alignment.indent,
                },
                sort_keys=True,
                default=str,
            ),
            protection=json.dumps(
                {"locked": cell.protection.locked, "hidden": cell.protection.hidden},
                sort_keys=True,
                default=str,
            ),
        )

    @staticmethod
    def _color_summary(color: Any) -> dict[str, ScalarValue] | None:
        if color is None:
            return None
        color_type = getattr(color, "type", None)
        value = getattr(color, "value", None)
        tint = getattr(color, "tint", 0.0)
        return {
            "type": None if color_type is None else str(color_type),
            "value": WorkbookParser._normalise_value(value),
            "tint": WorkbookParser._normalise_value(tint),
        }

    def _preflight_sheet_xml(
        self,
        archive: zipfile.ZipFile,
        infos: list[zipfile.ZipInfo],
    ) -> None:
        materialized_cells = 0
        merged_expansion = 0
        try:
            worksheet_parts = self._worksheet_parts(archive, infos)
            for info in infos:
                if self._normalise_archive_name(info.filename) not in worksheet_parts:
                    continue
                with archive.open(info) as stream:
                    saw_root = False
                    for event, element in ElementTree.iterparse(
                        stream,
                        events=("start", "end"),
                    ):
                        tag = element.tag.rsplit("}", 1)[-1]
                        if not saw_root and event == "start":
                            saw_root = True
                            if tag != "worksheet":
                                raise WorkbookParseError(
                                    f"Worksheet part {info.filename!r} has an invalid root element"
                                )
                        if event != "end":
                            continue
                        if tag == "c":
                            reference = element.attrib.get("r", "")
                            min_column, min_row, max_column, max_row = range_boundaries(reference)
                            if (
                                None in {min_column, min_row, max_column, max_row}
                                or int(min_column) < 1
                                or int(max_column) > 16_384
                                or int(min_row) < 1
                                or int(max_row) > 1_048_576
                            ):
                                raise WorkbookParseError(
                                    f"Workbook contains an out-of-bounds cell: {reference!r}"
                                )
                            materialized_cells += 1
                        elif tag == "mergeCell":
                            reference = element.attrib.get("ref", "")
                            min_column, min_row, max_column, max_row = range_boundaries(reference)
                            if None in {min_column, min_row, max_column, max_row}:
                                raise WorkbookParseError(
                                    f"Workbook contains an invalid merged range: {reference!r}"
                                )
                            if (
                                int(min_column) < 1
                                or int(max_column) > 16_384
                                or int(min_row) < 1
                                or int(max_row) > 1_048_576
                            ):
                                raise WorkbookParseError(
                                    f"Merged range {reference!r} falls outside Excel limits"
                                )
                            area = (int(max_column) - int(min_column) + 1) * (
                                int(max_row) - int(min_row) + 1
                            )
                            if area > self.max_merged_cells:
                                raise WorkbookParseError(
                                    f"Merged range {reference!r} expands to {area:,} cells; "
                                    f"the configured limit is {self.max_merged_cells:,}"
                                )
                            merged_expansion += max(0, area - 1)
                        element.clear()
                        if materialized_cells + merged_expansion > self.max_cells:
                            raise WorkbookParseError(
                                "Workbook contains more than the configured "
                                f"{self.max_cells:,} cells after merged-range expansion"
                            )
        except WorkbookParseError:
            raise
        except (ElementTree.ParseError, TypeError, ValueError) as exc:
            raise WorkbookParseError(f"Unable to preflight worksheet XML safely: {exc}") from exc

    def _worksheet_parts(
        self,
        archive: zipfile.ZipFile,
        infos: list[zipfile.ZipInfo],
    ) -> set[str]:
        """Resolve worksheet parts by OPC metadata, independent of filename extension."""

        names = {self._normalise_archive_name(info.filename) for info in infos}
        parts: set[str] = set()

        content_types_name = "[Content_Types].xml"
        if content_types_name in names:
            root = ElementTree.fromstring(archive.read(content_types_name))
            worksheet_extensions: set[str] = set()
            for element in root:
                tag = element.tag.rsplit("}", 1)[-1]
                content_type = element.attrib.get("ContentType", "").lower()
                if not content_type.endswith(_WORKSHEET_CONTENT_TYPE):
                    continue
                if tag == "Override":
                    parts.add(
                        self._normalise_archive_name(element.attrib.get("PartName", ""))
                    )
                elif tag == "Default":
                    extension = element.attrib.get("Extension", "").lower().lstrip(".")
                    if extension:
                        worksheet_extensions.add(extension)
            for name in names:
                if name.rsplit(".", 1)[-1].lower() in worksheet_extensions:
                    parts.add(name)

        for info in infos:
            relationship_name = self._normalise_archive_name(info.filename)
            if not relationship_name.endswith(".rels"):
                continue
            root = ElementTree.fromstring(archive.read(info))
            source_part = self._relationship_source_part(relationship_name)
            if source_part is None:
                continue
            source_directory = posixpath.dirname(source_part)
            for relationship in root:
                relationship_type = relationship.attrib.get("Type", "").lower()
                if not relationship_type.endswith(_WORKSHEET_RELATIONSHIP):
                    continue
                if relationship.attrib.get("TargetMode", "").lower() == "external":
                    continue
                target = relationship.attrib.get("Target", "")
                resolved = self._normalise_archive_name(
                    posixpath.join(source_directory, target)
                )
                parts.add(resolved)

        return {part for part in parts if part in names}

    @staticmethod
    def _relationship_source_part(relationship_name: str) -> str | None:
        if relationship_name == "_rels/.rels":
            return ""
        marker = "/_rels/"
        if marker not in relationship_name:
            return None
        prefix, filename = relationship_name.rsplit(marker, 1)
        if not filename.endswith(".rels"):
            return None
        return posixpath.join(prefix, filename[: -len(".rels")])

    @staticmethod
    def _normalise_archive_name(name: str) -> str:
        normalised = posixpath.normpath(name.replace("\\", "/").lstrip("/"))
        if normalised == "." or normalised.startswith("../"):
            return ""
        return normalised

    @staticmethod
    def _hidden_rows(dimensions: Iterable[tuple[int, Any]]) -> list[int]:
        hidden: list[int] = []
        for row_index, dimension in dimensions:
            if not bool(dimension.hidden):
                continue
            if row_index < 1 or row_index > 1_048_576:
                raise WorkbookParseError(
                    f"Hidden row {row_index} is outside Excel worksheet limits"
                )
            hidden.append(row_index)
        return sorted(hidden)

    @staticmethod
    def _hidden_columns(dimensions: Iterable[Any]) -> list[str]:
        hidden: set[str] = set()
        for dimension in dimensions:
            if not bool(dimension.hidden):
                continue
            minimum = int(dimension.min or 0)
            maximum = int(dimension.max or minimum)
            if minimum < 1 or maximum < minimum or maximum > 16_384:
                raise WorkbookParseError(
                    f"Hidden column range {minimum}:{maximum} is outside Excel limits"
                )
            hidden.update(get_column_letter(index) for index in range(minimum, maximum + 1))
        return sorted(hidden)

    @staticmethod
    def _data_validations(worksheet: Any) -> list[DataValidationInfo]:
        container = getattr(worksheet, "data_validations", None)
        validations = getattr(container, "dataValidation", []) if container else []
        result = [
            DataValidationInfo(
                ranges=str(item.sqref),
                validation_type=item.type,
                operator=item.operator,
                formula1=None if item.formula1 is None else str(item.formula1),
                formula2=None if item.formula2 is None else str(item.formula2),
                allow_blank=bool(item.allowBlank),
                show_error_message=bool(item.showErrorMessage),
            )
            for item in validations
        ]
        return sorted(result, key=lambda item: (item.ranges, item.validation_type or ""))

    @staticmethod
    def _named_ranges(workbook: Any) -> list[NamedRangeInfo]:
        values = getattr(workbook.defined_names, "values", None)
        definitions = values() if callable(values) else []
        result = [
            NamedRangeInfo(
                name=str(item.name),
                value=str(item.attr_text or ""),
                kind=getattr(item, "type", None),
                local_sheet_id=getattr(item, "localSheetId", None),
                hidden=getattr(item, "hidden", None),
            )
            for item in definitions
        ]
        return sorted(result, key=lambda item: (item.name, item.local_sheet_id or -1))

    @staticmethod
    def _external_links(workbook: Any, sheets: list[SheetSnapshot]) -> list[str]:
        links: set[str] = set()
        for link in getattr(workbook, "_external_links", []):
            file_link = getattr(link, "file_link", None)
            target = getattr(file_link, "Target", None)
            links.add(str(target or link))
        for sheet in sheets:
            for cell in sheet.cells.values():
                if cell.formula:
                    links.update(_EXTERNAL_BOOK_RE.findall(cell.formula))
        return sorted(links)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
